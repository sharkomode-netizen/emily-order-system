"""
Feedback Processor — 自动分析用户反馈，提取规则，验证并应用。
用户提交反馈后由后台线程触发。
"""
import json
import os
import re
import time
import logging
import traceback
import urllib.request
import urllib.parse
from datetime import datetime

logger = logging.getLogger(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FEEDBACK_FILE = os.path.join(BASE_DIR, 'feedback.json')
FEEDBACK_FILES_DIR = os.path.join(BASE_DIR, 'feedback_files')
RULES_FILE = os.path.join(BASE_DIR, 'correction_rules.json')

# WhatsApp notification (CallMeBot)
WHATSAPP_PHONE = os.environ.get('WHATSAPP_PHONE', '+8613528444234')
WHATSAPP_APIKEY = os.environ.get('WHATSAPP_APIKEY', '')

# Rate limiting
MAX_PROCESS_PER_HOUR = 5
_process_timestamps = []

# Dangerous keywords (shared with review_feedback.py)
DANGER_KEYWORDS = [
    'drop table', 'delete all', 'rm -rf', 'shutdown', 'exec(', 'eval(',
    'import os', '__import__', 'system(', 'subprocess', 'sql injection',
    '<script', 'javascript:', 'onerror', 'onclick',
]


def _is_dangerous(msg):
    lower = msg.lower()
    return any(kw in lower for kw in DANGER_KEYWORDS)


def _notify_whatsapp(message):
    """Send WhatsApp notification via CallMeBot."""
    if not WHATSAPP_APIKEY:
        logger.warning("No WHATSAPP_APIKEY, skipping WhatsApp notification")
        return False
    try:
        text = urllib.parse.quote_plus(message[:1000])
        phone = urllib.parse.quote_plus(WHATSAPP_PHONE)
        url = f"https://api.callmebot.com/whatsapp.php?phone={phone}&text={text}&apikey={WHATSAPP_APIKEY}"
        req = urllib.request.Request(url, method='GET')
        with urllib.request.urlopen(req, timeout=10) as resp:
            logger.info(f"WhatsApp notification sent: {resp.status}")
            return resp.status == 200
    except Exception as e:
        logger.error(f"WhatsApp notification failed: {e}")
        return False


def _rate_limited():
    """Check if we've exceeded processing rate limit."""
    global _process_timestamps
    now = time.time()
    _process_timestamps = [t for t in _process_timestamps if now - t < 3600]
    return len(_process_timestamps) >= MAX_PROCESS_PER_HOUR


def _call_ai(prompt, timeout=60):
    """Call Claude API for feedback analysis."""
    api_key = os.environ.get('ANTHROPIC_API_KEY', '')
    if not api_key:
        logger.warning("No ANTHROPIC_API_KEY, skipping AI analysis")
        return None
    try:
        from anthropic import Anthropic
        client = Anthropic(api_key=api_key, timeout=timeout)
        response = client.messages.create(
            model='claude-sonnet-4-20250514',
            max_tokens=2000,
            messages=[{'role': 'user', 'content': prompt}],
        )
        return response.content[0].text
    except Exception as e:
        logger.error(f"AI call failed: {e}")
        return None


def _parse_ai_json(text):
    """Extract JSON from AI response."""
    if not text:
        return None
    start = text.find('{')
    end = text.rfind('}')
    if start >= 0 and end > start:
        try:
            return json.loads(text[start:end + 1])
        except json.JSONDecodeError:
            # Try finding a JSON array
            start = text.find('[')
            end = text.rfind(']')
            if start >= 0 and end > start:
                try:
                    return json.loads(text[start:end + 1])
                except json.JSONDecodeError:
                    pass
    return None


def _load_feedback():
    if not os.path.exists(FEEDBACK_FILE):
        return []
    try:
        with open(FEEDBACK_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return []


def _save_feedback(feedbacks):
    with open(FEEDBACK_FILE, 'w', encoding='utf-8') as f:
        json.dump(feedbacks, f, ensure_ascii=False, indent=2)


def _read_excel_as_text(filepath, max_rows=50):
    """Read Excel file and return text representation for AI analysis."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        lines = []
        for ws in wb.worksheets:
            lines.append(f"=== Sheet: {ws.title} ===")
            for r in range(1, min(ws.max_row + 1, max_rows)):
                row_vals = []
                for c in range(1, min(ws.max_column + 1, 30)):
                    v = ws.cell(r, c).value
                    if v is not None:
                        row_vals.append(str(v)[:50])
                if any(row_vals):
                    lines.append(f"Row {r}: {' | '.join(v for v in row_vals if v)}")
        return '\n'.join(lines)
    except Exception as e:
        logger.error(f"Failed to read Excel: {e}")
        return None


def process_single(fb_id):
    """Process a single feedback item — main entry point for background thread."""
    if _rate_limited():
        logger.info(f"Rate limited, skipping feedback #{fb_id}")
        return

    feedbacks = _load_feedback()
    fb = None
    for f in feedbacks:
        if f.get('id') == fb_id:
            fb = f
            break

    if not fb:
        logger.warning(f"Feedback #{fb_id} not found")
        return

    if fb.get('status') != 'pending':
        return

    msg = fb.get('message', '')
    category = fb.get('category', '')

    # Safety check
    if _is_dangerous(msg):
        fb['status'] = 'rejected'
        fb['process_log'] = 'Dangerous content detected'
        _save_feedback(feedbacks)
        return

    # Skip too short
    if len(msg.strip()) < 3 and category != 'training':
        fb['status'] = 'skipped'
        fb['process_log'] = 'Too short'
        _save_feedback(feedbacks)
        return

    _process_timestamps.append(time.time())

    try:
        if category == 'training':
            rules = _process_training_data(fb)
        else:
            rules = _process_text_feedback(fb)

        if rules:
            from rule_engine import save_rule
            saved_count = 0
            pending_review = []
            for rule in rules:
                if save_rule(rule):
                    saved_count += 1
                    if not rule.get('verified'):
                        pending_review.append(rule)

            # Notify about low-confidence rules needing review
            if pending_review:
                descriptions = '\n'.join(
                    f"- [{r['confidence']:.0%}] {r.get('description', '')[:80]}"
                    for r in pending_review
                )
                _notify_whatsapp(
                    f"[EMILY] {len(pending_review)} 条规则待审核 (置信度<70%):\n"
                    f"{descriptions}\n"
                    f"反馈#{fb_id}: {msg[:100]}"
                )

            fb['status'] = 'processed'
            fb['process_log'] = f'Generated {len(rules)} rules, saved {saved_count} new ({len(pending_review)} pending review)'
            fb['rules_generated'] = [r.get('id') for r in rules]
        else:
            fb['status'] = 'processed'
            fb['process_log'] = 'No actionable rules extracted'

    except Exception as e:
        fb['status'] = 'error'
        fb['process_log'] = f'Processing error: {str(e)}'
        logger.error(f"Feedback #{fb_id} processing failed: {e}", exc_info=True)

    _save_feedback(feedbacks)
    logger.info(f"Feedback #{fb_id} processed: {fb.get('process_log')}")


def _process_training_data(fb):
    """Analyze training data pair (source file → correct target file)."""
    files = fb.get('files', [])
    source_file = None
    target_file = None

    for f in files:
        path = os.path.join(FEEDBACK_FILES_DIR, f.get('path', ''))
        if not os.path.exists(path):
            continue
        if f.get('role') == 'source':
            source_file = path
        elif f.get('role') == 'target':
            target_file = path

    if not source_file or not target_file:
        return []

    # Read both files
    source_text = _read_excel_as_text(source_file)
    target_text = _read_excel_as_text(target_file)

    if not source_text or not target_text:
        return []

    training_type = fb.get('training_type', 'po_to_pi')
    user_notes = fb.get('message', '')

    prompt = f"""你是一个数据分析专家。比较以下两个文件的差异，提取出系统性的规则。

这是一个 {training_type} 转换。用户认为 TARGET 是正确的输出。

SOURCE (系统输入):
{source_text[:3000]}

TARGET (正确输出):
{target_text[:3000]}

用户说明: {user_notes[:500]}

请找出系统性差异（不是一次性错误），输出JSON格式:
{{
  "differences": [
    {{
      "field": "哪个字段有问题",
      "source_pattern": "源文件中的模式",
      "target_pattern": "目标文件中的正确模式",
      "rule_description": "用英文描述规则，如 'strip leading digit prefix from style code'",
      "rule_type": "field_mapping 或 value_correction 或 format_override 或 prompt_patch",
      "transform": "round/strip_prefix/replace/set_default/uppercase/lowercase/number_format 之一",
      "transform_params": {{}},
      "confidence": 0.0到1.0
    }}
  ]
}}

只输出高置信度(>0.7)的系统性规则。如果没有发现系统性差异，输出空的differences数组。"""

    result = _call_ai(prompt)
    if not result:
        return []

    parsed = _parse_ai_json(result)
    if not parsed:
        return []

    differences = parsed if isinstance(parsed, list) else parsed.get('differences', [])
    rules = []

    for diff in differences:
        confidence = diff.get('confidence', 0)

        transform = diff.get('transform', '')
        if transform not in ('round', 'strip_prefix', 'replace', 'set_default',
                            'uppercase', 'lowercase', 'number_format'):
            # For prompt_patch type, transform is not needed
            if diff.get('rule_type') != 'prompt_patch':
                continue

        rule_id = f"auto_{fb.get('id')}_{len(rules) + 1}_{int(time.time())}"
        auto_verified = confidence >= 0.7

        rule = {
            'id': rule_id,
            'type': diff.get('rule_type', 'value_correction'),
            'scope': 'parse',
            'conversion': training_type,
            'description': diff.get('rule_description', ''),
            'condition': {},
            'action': {
                'field': diff.get('field', ''),
                'transform': transform,
            },
            'confidence': confidence,
            'source_feedback_id': fb.get('id'),
            'created': datetime.now().isoformat(),
            'verified': auto_verified,
            'enabled': auto_verified,
        }

        # Add transform params
        params = diff.get('transform_params', {})
        if params:
            rule['action'].update(params)

        # For prompt_patch type
        if diff.get('rule_type') == 'prompt_patch':
            rule['scope'] = 'parse'
            rule['action'] = {
                'append_instruction': diff.get('rule_description', ''),
            }

        rules.append(rule)

    return rules


def _process_text_feedback(fb):
    """Analyze text feedback and propose rules."""
    msg = fb.get('message', '')
    category = fb.get('category', '')

    # Only process data/format/feature categories
    if category not in ('data', 'format', 'feature'):
        return []

    prompt = f"""你是一个系统优化专家。分析以下用户反馈，判断是否可以提取出自动化规则。

系统是一个鞋类订单转换工具，有5个功能:
1. PO→PI (采购订单→形式发票)
2. PI→Packing List (形式发票→装箱单)
3. PI→CI (形式发票→商业发票)
4. 手写材料→生产指令单
5. 报价单→COG (成本表)

用户反馈:
分类: {category}
内容: {msg[:1000]}

如果可以提取规则，输出JSON:
{{
  "actionable": true,
  "rules": [
    {{
      "rule_type": "field_mapping/value_correction/format_override/prompt_patch",
      "conversion": "po_to_pi/pi_to_packing/pi_to_ci/materials_to_production/quotation_to_cog/all",
      "description": "规则英文描述",
      "transform": "round/strip_prefix/replace/set_default/uppercase/lowercase/number_format",
      "field": "affected field name",
      "transform_params": {{}},
      "confidence": 0.0-1.0
    }}
  ]
}}

如果反馈不够具体或不可操作，输出:
{{"actionable": false, "reason": "why"}}"""

    result = _call_ai(prompt)
    if not result:
        return []

    parsed = _parse_ai_json(result)
    if not parsed or not parsed.get('actionable'):
        return []

    rules = []
    for r in parsed.get('rules', []):
        confidence = r.get('confidence', 0)

        rule_id = f"auto_{fb.get('id')}_{len(rules) + 1}_{int(time.time())}"
        transform = r.get('transform', '')
        auto_verified = confidence >= 0.7

        rule = {
            'id': rule_id,
            'type': r.get('rule_type', 'value_correction'),
            'scope': 'parse',
            'conversion': r.get('conversion', 'all'),
            'description': r.get('description', ''),
            'condition': {},
            'action': {
                'field': r.get('field', ''),
                'transform': transform,
            },
            'confidence': confidence,
            'source_feedback_id': fb.get('id'),
            'created': datetime.now().isoformat(),
            'verified': auto_verified,
            'enabled': auto_verified,
        }

        params = r.get('transform_params', {})
        if params:
            rule['action'].update(params)

        if r.get('rule_type') == 'prompt_patch':
            rule['action'] = {'append_instruction': r.get('description', '')}

        rules.append(rule)

    return rules


def process_all_pending():
    """Process all pending actionable feedback. Called on startup or by cron."""
    feedbacks = _load_feedback()
    pending = [f for f in feedbacks if f.get('status') == 'pending']

    processed = 0
    for fb in pending:
        if _rate_limited():
            logger.info(f"Rate limited after processing {processed} items")
            break

        category = fb.get('category', '')
        msg = fb.get('message', '')

        # Only process actionable categories
        if category not in ('data', 'format', 'feature', 'training'):
            continue
        if _is_dangerous(msg):
            continue
        if len(msg.strip()) < 3 and category != 'training':
            continue

        process_single(fb['id'])
        processed += 1

    logger.info(f"Processed {processed} pending feedback items")
    return processed


if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO)
    n = process_all_pending()
    print(f"Processed {n} items")
