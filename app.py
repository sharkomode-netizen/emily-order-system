"""
EMILY 订单管理系统 v2.1
- PO → PI 转换：上传客户PO(PDF/Excel .xlsx/.xls/图片)，自动生成EMILY格式PI
- 手写原材料 → 生产指令单：上传手写采购需求表照片，AI识别后生成生产指令单
- 支持多端口运行，可部署到云服务器
"""

import os
import json
import base64
import re
import subprocess
import sys
import io
import traceback
import time
from datetime import datetime
from pathlib import Path
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify, Response
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import pdfplumber

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'emily-order-system-2026')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app.config['UPLOAD_FOLDER'] = os.path.join(BASE_DIR, 'uploads')
app.config['OUTPUT_FOLDER'] = os.path.join(BASE_DIR, 'output')
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024
FEEDBACK_FILE = os.path.join(BASE_DIR, 'feedback.json')

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

# Claude CLI 路径（本地开发）或 API key（云端部署）
CLAUDE_CLI = os.environ.get('CLAUDE_CLI', '/Users/danny/.nvm/versions/node/v24.13.1/bin/claude')
ANTHROPIC_API_KEY = os.environ.get('ANTHROPIC_API_KEY', '')


def _parse_ai_json(text):
    """从AI返回文本中提取JSON，解析失败返回None"""
    json_match = re.search(r'\{[\s\S]*\}', text)
    if json_match:
        try:
            return json.loads(json_match.group())
        except json.JSONDecodeError:
            return None
    return None


def safe_filename(filename):
    """Sanitize filename to prevent path traversal"""
    filename = os.path.basename(filename)
    filename = re.sub(r'[^\w\s\-\.\(\)]', '_', filename)
    return filename


def cleanup_uploads(max_age_hours=24):
    """Clean up old upload files"""
    now = time.time()
    upload_dir = app.config['UPLOAD_FOLDER']
    for f in os.listdir(upload_dir):
        fpath = os.path.join(upload_dir, f)
        if os.path.isfile(fpath) and (now - os.path.getmtime(fpath)) > max_age_hours * 3600:
            try:
                os.remove(fpath)
            except OSError:
                pass


# ============================================================
# AI 调用层：优先 API，回退 CLI
# ============================================================
def call_ai(prompt, images=None, timeout=180):
    if ANTHROPIC_API_KEY:
        return _call_api(prompt, images, timeout)
    else:
        return _call_cli(prompt, images, timeout)


def _call_api(prompt, images=None, timeout=180):
    from anthropic import Anthropic
    client = Anthropic(api_key=ANTHROPIC_API_KEY, timeout=timeout)

    content = []
    if images:
        for img_data in images:
            content.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": img_data['media_type'],
                    "data": img_data['data'],
                }
            })
    content.append({"type": "text", "text": prompt})

    try:
        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=8192,
            messages=[{"role": "user", "content": content}]
        )
        return response.content[0].text
    except Exception as e:
        raise RuntimeError(f"AI API 调用失败: {str(e)[:300]}")


def _call_cli(prompt, images=None, timeout=180):
    full_prompt = prompt
    if images and isinstance(images, list) and all(isinstance(i, dict) and 'path' in i for i in images):
        paths = "\n".join(f"- {img['path']}" for img in images)
        full_prompt = f"请先用Read工具读取以下图片文件，然后根据图片内容回答：\n{paths}\n\n{prompt}"

    env = os.environ.copy()
    env.pop('CLAUDECODE', None)

    try:
        result = subprocess.run(
            [CLAUDE_CLI, '-p', '--output-format', 'json'],
            input=full_prompt,
            capture_output=True,
            text=True,
            timeout=timeout,
            env=env,
        )
    except subprocess.TimeoutExpired:
        raise RuntimeError(f"AI 处理超时（{timeout}秒），请重试或减少上传文件数量")

    if result.returncode != 0:
        raise RuntimeError(f"AI 调用失败: {result.stderr[:200]}")

    for line in result.stdout.strip().split('\n'):
        line = line.strip()
        if line.startswith('{'):
            try:
                data = json.loads(line)
                if data.get('type') == 'result':
                    return data.get('result', '')
            except json.JSONDecodeError:
                continue
    return result.stdout


# ============================================================
# EMILY 公司信息
# ============================================================
EMILY_INFO = {
    'company_en': 'EMILY HONG KONG LIMITED',
    'address': 'FLAT/RM 20, 8/F, YALE INDUSTRIAL CENTRE, 61-63 AU PUI WAN STREET, FO TAN, SHATIN, NT, HONG KONG',
    'tel': '+852 XXXX XXXX',
    'fax': '+852 XXXX XXXX',
    'bank_name': 'E.SUN COMMERCIAL BANK LTD., HONG KONG BRANCH',
    'bank_address': 'SUITE 2805, 28F, TOWER 6, THE GATEWAY, 9 CANTON ROAD, TSIMSHATSUI, KOWLOON, HONG KONG',
    'beneficiary': 'EMILY HONG KONG LIMITED',
    'beneficiary_address': 'FLAT/RM 20, 8/F, YALE INDUSTRIAL CENTRE, 61-63 AU PUI WAN STREET, FO TAN, SHATIN, NT, HONG KONG',
    'swift': 'ESUNHKHH',
    'account': '909441141068',
}


# ============================================================
# .xls → .xlsx 转换
# ============================================================
def convert_xls_to_xlsx(xls_path):
    """Convert old .xls format to .xlsx using xlrd + openpyxl"""
    import xlrd
    xls_wb = xlrd.open_workbook(xls_path)
    xls_ws = xls_wb.sheet_by_index(0)

    xlsx_wb = openpyxl.Workbook()
    xlsx_ws = xlsx_wb.active

    for r in range(xls_ws.nrows):
        for c in range(xls_ws.ncols):
            cell = xls_ws.cell(r, c)
            val = cell.value
            # xlrd date handling
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    val = xlrd.xldate_as_datetime(val, xls_wb.datemode)
                except Exception:
                    pass
            xlsx_ws.cell(row=r + 1, column=c + 1, value=val)

    # Handle merged cells
    for merge in xls_ws.merged_cells:
        r1, r2, c1, c2 = merge
        xlsx_ws.merge_cells(
            start_row=r1 + 1, start_column=c1 + 1,
            end_row=r2, end_column=c2
        )

    xlsx_path = xls_path.rsplit('.', 1)[0] + '_converted.xlsx'
    xlsx_wb.save(xlsx_path)
    return xlsx_path


# ============================================================
# PI 解析
# ============================================================
def parse_pi_excel(filepath):
    """解析PI Excel文件，支持多种客户格式。全部结构化解析，不调AI"""
    # .xls 转换
    if filepath.lower().endswith('.xls') and not filepath.lower().endswith('.xlsx'):
        filepath = convert_xls_to_xlsx(filepath)

    # 先尝试MAYORAL格式（行内多尺码）
    try:
        result = _parse_pi_excel_structured(filepath)
        if result and result.get('items'):
            return result
    except Exception:
        pass

    # 再尝试通用格式（一行一尺码，如PROTETIKA）
    try:
        result = _parse_pi_generic(filepath)
        if result and result.get('items'):
            return result
    except Exception:
        traceback.print_exc()

    # 都失败了才用AI（最慢）
    return _parse_pi_excel_ai(filepath)


def _parse_pi_generic(filepath):
    """通用PI解析器：自动检测列头，支持一行一尺码格式。扫描所有sheet找最佳数据源"""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # 扫描所有sheet，找到含列头的最佳sheet
    header_keywords = {
        'article': ['article', 'art.', 'style', 'item no', 'model', 'varegrundkode'],
        'sku': ['varenummer', 'sku', 'item code', 'product code'],
        'description': ['varebeskrivelse', 'description', 'desc', 'product name'],
        'color_code': ['farvekode', 'color code', 'colour code', 'col code'],
        'color_name': ['colour name', 'color name', 'farvenavn'],
        'size': ['size', 'sizes', 'størrelseskode'],
        'qty': ['qty', 'quantity', 'pairs', 'pcs', 'antal'],
        'price': ['price', 'unit price', 'unit_price'],
        'amount': ['amount', 'total amount', 'total'],
    }

    best_sheet = None
    best_header_row = None
    best_col_map = {}
    best_matches = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for r in range(1, min(ws.max_row + 1, 30)):
            row_text = {}
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v:
                    row_text[c] = str(v).strip().lower()

            matches = 0
            tmp_map = {}
            for role, keywords in header_keywords.items():
                for c, txt in row_text.items():
                    if any(kw in txt for kw in keywords):
                        if role not in tmp_map:
                            tmp_map[role] = c
                            matches += 1
                        break

            if matches > best_matches:
                best_matches = matches
                best_sheet = ws
                best_header_row = r
                best_col_map = tmp_map

    if best_matches < 3 or not best_sheet:
        return None

    ws = best_sheet
    header_row = best_header_row
    col_map = best_col_map

    # === 2. 提取头部信息（扫描所有sheet的前面几行） ===
    order_no = ''
    date_str = ''
    customer = ''
    customer_address = ''
    customer_vat = ''
    bank_info = {}
    currency = 'USD'
    total_qty_header = 0
    total_amt_header = 0

    for sn in wb.sheetnames:
        sws = wb[sn]
        for r in range(1, min(sws.max_row + 1, 25)):
            for c in range(1, min(sws.max_column + 1, 10)):
                v = str(sws.cell(r, c).value or '').strip()
                if not v:
                    continue
                vl = v.lower()

                # PO号
                if not order_no:
                    m = re.search(r'(PO\d+)', v)
                    if m:
                        order_no = m.group(1)
                    elif any(k in vl for k in ['order n', 'order no', 'order nr', 'po no', 'po nr', 'dokumentnummer']):
                        nv = sws.cell(r, c + 1).value
                        if nv:
                            order_no = str(nv).strip()

                # 客户名（从Vendor/Ship To列 — 注意EMILY是seller不是buyer）
                if not customer and ('vendor' in vl or 'ship to' in vl) and 'no.' not in vl and 'number' not in vl:
                    nv = sws.cell(r, c + 1).value
                    if nv:
                        name_candidate = str(nv).strip()
                        ncl = name_candidate.lower()
                        # 排除EMILY自身、DIRECT、以及列头名
                        if ('emily' not in ncl and name_candidate.upper() != 'DIRECT'
                                and 'kunde' not in ncl and 'leverandør' not in ncl
                                and 'customer' not in ncl and 'supplier' not in ncl):
                            customer = name_candidate

                # ETD / 日期
                if 'etd' in vl or 'date' in vl:
                    nv = sws.cell(r, c + 1).value
                    if nv and not date_str:
                        if hasattr(nv, 'strftime'):
                            date_str = nv.strftime('%d-%b-%Y')
                        else:
                            date_str = str(nv).strip()

                # 总数（从header提取，用于校验）
                if 'total qty' in vl or 'total pair' in vl:
                    nv = sws.cell(r, c + 1).value
                    if nv:
                        try:
                            total_qty_header = int(float(nv))
                        except (ValueError, TypeError):
                            pass
                if 'total amount' in vl:
                    nv = sws.cell(r, c + 1).value
                    if nv:
                        try:
                            total_amt_header = float(nv)
                        except (ValueError, TypeError):
                            pass

                # 买家
                if 'buyer' in vl:
                    nv = sws.cell(r + 1, c).value
                    if nv and not customer:
                        customer = str(nv).strip()
                        addr = sws.cell(r + 2, c).value
                        if addr:
                            customer_address = str(addr).strip()

                # 客户名（从 Customer 列提取，排除 EMILY 自身）
                if not customer and ('customer' in vl):
                    nv = sws.cell(r, c + 1).value
                    if nv:
                        name_candidate = str(nv).strip()
                        if 'emily' not in name_candidate.lower() and name_candidate.upper() != 'DIRECT':
                            customer = name_candidate

                # VAT
                if 'vat' in vl:
                    nv = sws.cell(r, c + 1).value
                    if nv and not customer_vat:
                        customer_vat = str(nv).strip()

                # 银行
                if 'iban' in vl:
                    nv = sws.cell(r, c + 1).value
                    if nv:
                        bank_info['account'] = str(nv).strip()
                if 'bic' in vl or 'swift' in vl:
                    nv = sws.cell(r, c + 1).value
                    if nv:
                        bank_info['swift'] = str(nv).strip()
                if 'bank' in vl and 'detail' not in vl:
                    nv = sws.cell(r + 1, c).value
                    if nv and 'iban' not in str(nv).lower():
                        bank_info['bank_name'] = str(nv).strip()

                # 币种
                if 'eur' in vl:
                    currency = 'EUR'
                if 'usd' in vl:
                    currency = 'USD'

    # Check sub-header row for currency
    sub_row = header_row + 1
    for c in range(1, ws.max_column + 1):
        sv = str(ws.cell(sub_row, c).value or '').lower()
        if 'eur' in sv:
            currency = 'EUR'
        elif 'usd' in sv:
            currency = 'USD'

    # === 3. 读取数据行，按 article+color 分组 ===
    art_col = col_map.get('article', 1)
    sku_col = col_map.get('sku')
    color_name_col = col_map.get('color_name')
    desc_col = col_map.get('description')
    # For name_col: prefer color_name, fallback to description
    name_col = color_name_col or desc_col or 2
    color_code_col = col_map.get('color_code')
    size_col = col_map.get('size', 3)
    qty_col = col_map.get('qty', 5)
    price_col = col_map.get('price')
    amt_col = col_map.get('amount')

    # Detect Scandinavian/European format → default EUR
    if currency == 'USD':
        scandinavian_markers = ['dokumentnummer', 'varebeskrivelse', 'farvekode', 'størrelseskode', 'antal', 'varegrundkode']
        header_text = ' '.join(str(ws.cell(best_header_row, c).value or '').lower() for c in range(1, ws.max_column + 1))
        if any(m in header_text for m in scandinavian_markers):
            currency = 'EUR'

    # Skip sub-header rows
    start_row = header_row + 1
    sv = str(ws.cell(start_row, qty_col).value or '').lower()
    if sv in ('pairs', 'pcs', 'qty', 'code', ''):
        start_row += 1

    groups = {}  # key -> {sizes, price, desc, ...}
    current_article = ''
    current_name = ''
    current_sku = ''
    current_cc = ''
    all_sizes = set()

    current_desc = ''
    current_cn = ''

    for r in range(start_row, ws.max_row + 1):
        art_v = ws.cell(r, art_col).value
        name_v = ws.cell(r, name_col).value
        size_v = ws.cell(r, size_col).value
        qty_v = ws.cell(r, qty_col).value
        price_v = ws.cell(r, price_col).value if price_col else None
        sku_v = ws.cell(r, sku_col).value if sku_col else None
        cc_v = ws.cell(r, color_code_col).value if color_code_col else None
        # Read color_name and description separately if both columns exist
        cn_v = ws.cell(r, color_name_col).value if color_name_col else None
        desc_v = ws.cell(r, desc_col).value if desc_col else None

        # 更新当前 article/name/sku/color_code
        if art_v:
            current_article = str(art_v).strip()
        if sku_v:
            current_sku = str(sku_v).strip()
        if cc_v:
            current_cc = str(cc_v).strip()
        if cn_v:
            current_cn = str(cn_v).strip()
        if desc_v:
            current_desc = str(desc_v).strip()
        if name_v:
            name_str = str(name_v).strip()
            # 跳过 "sizes 19-26:" 这种描述行
            if re.match(r'^sizes?\s+\d', name_str, re.IGNORECASE):
                pass
            else:
                current_name = name_str

        # 需要有尺码和数量
        if not size_v or not qty_v:
            if qty_v and not size_v and not art_v:
                continue
            if name_v and 'total' in str(name_v).lower():
                continue
            continue

        # 解析尺码
        size_str = str(size_v).strip().rstrip('#')
        try:
            size_num = int(float(size_str))
        except (ValueError, TypeError):
            continue

        if not (10 <= size_num <= 50):
            continue

        # 解析数量
        try:
            qty = int(float(qty_v))
        except (ValueError, TypeError):
            continue
        if qty <= 0:
            continue

        # 解析价格
        price = 0
        if price_v:
            try:
                price = float(price_v)
            except (ValueError, TypeError):
                pass

        # 解析行金额
        line_amount = 0
        if amt_col:
            amt_rv = ws.cell(r, amt_col).value
            if amt_rv:
                try:
                    line_amount = float(amt_rv)
                except (ValueError, TypeError):
                    pass
        if line_amount == 0:
            line_amount = price * qty

        # 分组 key：用 sku 或 article+name，加上 price 以支持分段单价
        # （同款式不同尺码段可能有不同单价）
        price_key = round(price, 2) if price > 0 else 0
        if sku_col and current_sku:
            key = (current_sku, price_key)
        else:
            key = (current_article, current_name, price_key)

        if key not in groups:
            groups[key] = {
                'article': current_article,
                'sku': current_sku if sku_col else '',
                'name': current_name,
                'description': current_desc if desc_col else '',
                'color_name_raw': current_cn if color_name_col else '',
                'color_code': current_cc if color_code_col else '',
                'price': price,
                'sizes': {},
                'total_amount': 0,
            }
        if price > 0 and groups[key]['price'] == 0:
            groups[key]['price'] = price
        groups[key]['sizes'][str(size_num)] = groups[key]['sizes'].get(str(size_num), 0) + qty
        groups[key]['total_amount'] += line_amount
        all_sizes.add(size_num)

    if not groups:
        return None

    # === 4. 构建 items ===
    items = []
    for key, g in groups.items():
        name = g['name']
        color_code = g.get('color_code', '')
        # Use dedicated color_name column if available, otherwise fallback to name
        color_name = g.get('color_name_raw', '') or name
        desc = g.get('description', '') or name

        # 尝试从 name 拆分 color_code（only if no dedicated columns）
        if not color_code and not g.get('color_name_raw'):
            m = re.match(r'^(\d+)\s+(.+)$', name)
            if m:
                color_code = m.group(1)
                color_name = m.group(2)

        total_qty = sum(g['sizes'].values())
        grp_amount = g['total_amount']
        avg_price = round(grp_amount / total_qty, 2) if total_qty > 0 and grp_amount > 0 else g['price']

        sku = g.get('sku', '')
        style = sku if sku else ''

        items.append({
            'style_code': g['article'],
            'style': style,
            'color_code': color_code,
            'color_name': color_name,
            'color_full': f"{color_code} {color_name}".strip() if color_code else color_name,
            'description': f"{g['article']} {desc}".strip() if desc else f"{g['article']} {color_name}",
            'price': avg_price,
            'pieces': total_qty,
            'sizes': g['sizes'],
            '_line_amount': grp_amount,
        })

    # === 5. 提取尾部条款 ===
    terms = {}
    for r in range(max(ws.max_row - 20, header_row), ws.max_row + 1):
        v = str(ws.cell(r, 1).value or '').strip()
        if not v:
            continue
        vl = v.lower()
        if 'payment' in vl:
            terms['payment'] = v
            m = re.search(r':\s*(.+)', v)
            if m:
                terms['payment'] = m.group(1).strip()
        elif 'shipment' in vl or 'latest' in vl:
            terms['shipment'] = v
            m = re.search(r':\s*(.+)', v)
            if m:
                terms['shipment_date'] = m.group(1).strip()
        elif 'fob' in vl or 'cif' in vl:
            terms['delivery'] = v
            m = re.search(r':\s*(.+)', v)
            if m:
                terms['delivery'] = m.group(1).strip()
        elif 'port' in vl:
            terms['port'] = v
        elif 'moq' in vl:
            terms['moq'] = v
        elif 'packing' in vl:
            terms['packing'] = v
        elif 'brand' in vl:
            terms['brand'] = v

    total_pieces = sum(it['pieces'] for it in items)
    total_amount = sum(it.get('_line_amount', 0) or (it['pieces'] * it['price']) for it in items)

    # 用 header 里的总数校验/替代（更精确）
    if total_qty_header > 0 and total_pieces != total_qty_header:
        # 如果差异小，用计算值；差异大说明漏了数据
        pass
    if total_amt_header > 0:
        total_amount = total_amt_header

    # 从文件名提取 PO 号和其他信息
    if not order_no:
        fname = os.path.basename(filepath)
        m = re.search(r'(PO\d+)', fname)
        if m:
            order_no = m.group(1)

    return {
        'order_no': order_no,
        'invoice_no': '',
        'date': date_str,
        'customer': customer,
        'customer_address': customer_address,
        'customer_vat': customer_vat,
        'items': items,
        'total_pieces': total_pieces,
        'total_amount': round(total_amount, 2),
        'currency': currency,
        'terms': terms,
        'bank_info': bank_info,
        'size_headers': sorted(all_sizes),
    }


def _extract_excel_text(filepath, max_rows=600):
    """提取Excel内容为文本，供AI解析。智能截断：保留头尾"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    all_lines = []
    for r in range(1, min(ws.max_row + 1, max_rows)):
        row_vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None:
                row_vals.append(str(v))
            else:
                row_vals.append('')
        line = '\t'.join(row_vals).rstrip('\t')
        if line.strip():
            all_lines.append(f"R{r}: {line}")

    text = '\n'.join(all_lines)
    # 如果太长，保留头部+尾部（条款通常在文件末尾）
    if len(text) > 20000:
        head = text[:14000]
        tail = text[-6000:]
        text = head + "\n\n... [MIDDLE ROWS OMITTED] ...\n\n" + tail
    return text


def _parse_pi_excel_ai(filepath):
    """用AI解析任意格式的PI Excel"""
    text = _extract_excel_text(filepath)

    prompt = f"""This is a shoe industry Proforma Invoice (PI) / Purchase Order from Excel.
EMILY HONG KONG LIMITED is the seller/manufacturer. The BUYER is the customer placing the order.

The format varies by customer. Some have one row per size, some have all sizes in columns.
Extract ALL data and return pure JSON:
{{
  "order_no": "order number",
  "invoice_no": "invoice/reference number",
  "date": "date string",
  "customer": "BUYER company name (NOT the seller EMILY)",
  "customer_address": "buyer address",
  "customer_vat": "buyer VAT number",
  "items": [
    {{
      "style_code": "article/style number",
      "style": "model name or number",
      "color_code": "color code if any",
      "color_name": "color name",
      "color_full": "full color string",
      "description": "product description",
      "price": unit_price_number,
      "pieces": total_qty_for_this_style_color,
      "sizes": {{"19": 10, "20": 10, "21": 20}}
    }}
  ],
  "total_pieces": total_pairs,
  "total_amount": total_amount,
  "currency": "EUR or USD",
  "terms": {{
    "brand": "brand if mentioned",
    "shipment": "shipment/delivery date line",
    "shipment_date": "date only",
    "port": "port info",
    "port_loading": "loading port",
    "port_destination": "destination",
    "delivery": "delivery terms (FOB/CIF)",
    "payment": "payment terms",
    "amount_in_words": "amount in words if present"
  }},
  "bank_info": {{
    "bank_name": "buyer's bank name",
    "bank_address": "bank address",
    "beneficiary_name": "beneficiary",
    "beneficiary_address": "beneficiary address",
    "swift": "SWIFT/BIC code",
    "account": "IBAN or account number"
  }},
  "size_headers": [19, 20, 21, ...]
}}

IMPORTANT:
- SPLIT PRICING: If the same style+color has DIFFERENT unit prices for different size ranges (e.g. sizes 19-25 at $9.30 and sizes 26-38 at $9.80), keep them as SEPARATE items in the array. Do NOT merge them. Each price tier = one item entry with only its sizes.
- If the same style+color has the SAME price across all sizes, combine into ONE item.
- size_headers should be the sorted list of ALL unique sizes across all items
- Split color into color_code (number) and color_name (text) when possible
- Currency might be EUR or USD - include it

Excel content (head and tail preserved, some middle rows may be omitted):
{text[:22000]}

Return ONLY valid JSON. If rows are omitted, still extract all visible items and terms from the tail section."""

    result = call_ai(prompt, timeout=300)
    return _parse_ai_json(result)


def _parse_pi_excel_structured(filepath):
    """结构化解析多尺码列PI Excel（MAYORAL格式、EMILY自有格式等）
    支持分段单价：同款不同尺码段不同价格保留为独立行"""

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # === 1. 扫描前30行提取头部信息 ===
    order_no = ''
    invoice_no = ''
    date_str = ''
    customer_name = ''
    customer_address = ''
    customer_vat = ''

    for r in range(1, min(ws.max_row + 1, 30)):
        for c in range(1, min(ws.max_column + 1, 15)):
            v = str(ws.cell(r, c).value or '').strip()
            if not v:
                continue

            # Order No / PO No
            if not order_no:
                m = re.search(r"(?:order\s*(?:no\.?|nr\.?)|Bisgaard'?s?\s*order\s*no\.?)\s*:?\s*(\w[\w-]*)", v, re.IGNORECASE)
                if m:
                    order_no = m.group(1)

            # Invoice No
            if not invoice_no:
                m = re.search(r'Invoice\s*No\.?\s*:?\s*([\w-]+)', v, re.IGNORECASE)
                if m:
                    invoice_no = m.group(1)

            # Date
            if not date_str:
                m = re.search(r'(?:Invoice\s*)?[Dd]ate\s*[:：]\s*(.+?)$', v)
                if m:
                    date_str = m.group(1).strip()

            # Bill-to / Messers 客户名
            if not customer_name:
                m = re.search(r'(?:Bill[\s-]*to|Messers?)\s*:?\s*$', v, re.IGNORECASE)
                if m:
                    # 客户名在下一行
                    nv = ws.cell(r + 1, c).value
                    if nv:
                        customer_name = str(nv).strip()
                        addr = ws.cell(r + 2, c).value
                        if addr:
                            customer_address = str(addr).strip()
                else:
                    m = re.search(r'Messers?:?\s*(.+?)(?:\s{3,}|Order)', v)
                    if m:
                        customer_name = m.group(1).strip()

            # VAT / CIF
            if not customer_vat:
                m = re.search(r'(?:C\.I\.F|VAT)\s*/?\s*(?:VAT)?\s*:?\s*([\w]+)', v, re.IGNORECASE)
                if m:
                    customer_vat = m.group(1).strip()

    # === 2. 找尺码头行和列头行 ===
    # 扫描所有行，找含有连续尺码数字（10-50）的行
    size_header_rows = {}
    col_header_row = 0
    item_col = 0
    color_col = 0
    desc_col = 0
    price_col = 0
    qty_col = 0
    amt_col_idx = 0
    code_col = 0

    for r in range(1, min(ws.max_row + 1, 30)):
        row_sizes = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            # 尺码可能是 "19#" 或纯数字19
            if v is not None:
                sv = str(v).strip().rstrip('#')
                try:
                    num = int(float(sv))
                    if 10 <= num <= 50:
                        row_sizes.append((c, num))
                except (ValueError, TypeError):
                    pass
        if len(row_sizes) >= 3:
            size_header_rows[r] = row_sizes

        # 检测列头行（含 Item/Article/Style/Color/Price 等关键词）
        row_text = ' '.join(str(ws.cell(r, c).value or '').lower() for c in range(1, ws.max_column + 1))
        if not col_header_row and ('item' in row_text or 'article' in row_text or 'style' in row_text):
            if 'color' in row_text or 'material' in row_text or 'desc' in row_text:
                col_header_row = r
                # 检测各列位置
                for c in range(1, ws.max_column + 1):
                    cv = str(ws.cell(r, c).value or '').lower().strip()
                    if not cv:
                        continue
                    if any(k in cv for k in ['item no', 'article', 'art.']):
                        item_col = c
                    elif any(k in cv for k in ['item code', 'sku', 'product code']):
                        code_col = c
                    elif any(k in cv for k in ['color', 'colour']):
                        color_col = c
                    elif any(k in cv for k in ['material', 'description', 'desc']):
                        desc_col = c
                    elif 'price' in cv:
                        price_col = c
                    elif cv in ('q\'ty', 'qty', 'quantity', 'pairs', 'pcs'):
                        qty_col = c
                    elif cv == 'amount' or 'amount' in cv:
                        amt_col_idx = c

    if not size_header_rows:
        return None

    # 取最早的尺码头行作为主尺码映射
    first_size_row = min(size_header_rows.keys())
    initial_sizes = size_header_rows[first_size_row]

    # 如果没找到列头行，用MAYORAL默认布局
    if not col_header_row:
        item_col = item_col or 2
        code_col = code_col or 3
        color_col = color_col or 4
        desc_col = desc_col or 5
        price_col = price_col or 6
    else:
        item_col = item_col or 1
        desc_col = desc_col or 4
        price_col = price_col or 0

    # 数据起始行：列头行或尺码头行之后
    data_start = max(col_header_row, first_size_row) + 1
    # 跳过子表头行（如包含 "PRS" "USD" 等）
    for skip_r in range(data_start, data_start + 3):
        sv = str(ws.cell(skip_r, price_col).value or '' if price_col else '').lower()
        if sv in ('', 'usd', 'eur', 'price', '(usd)', '(eur)'):
            data_start = skip_r + 1
        else:
            break

    # === 3. 逐行读取数据，按 style+color+price 分组 ===
    # 同款同色同价→合并尺码，不同价→保留为独立行（分段单价）
    groups = {}  # key=(item_no, color, price) -> group dict
    all_sizes_used = set()
    current_item_no = ''
    current_code = ''
    current_color = ''
    current_desc = ''

    sorted_size_rows = sorted(size_header_rows.keys())

    for r in range(data_start, ws.max_row + 1):
        # 更新当前字段（有些行只有尺码数据，item/color在上面的行）
        v_item = ws.cell(r, item_col).value if item_col else None
        v_code = ws.cell(r, code_col).value if code_col else None
        v_color = ws.cell(r, color_col).value if color_col else None
        v_desc = ws.cell(r, desc_col).value if desc_col else None
        v_price = ws.cell(r, price_col).value if price_col else None
        v_qty = ws.cell(r, qty_col).value if qty_col else None
        v_amt = ws.cell(r, amt_col_idx).value if amt_col_idx else None

        if v_item:
            current_item_no = str(int(v_item) if isinstance(v_item, float) else v_item).strip()
        if v_code:
            current_code = str(v_code).strip()
        if v_color:
            current_color = str(v_color).strip()
        if v_desc:
            current_desc = str(v_desc).strip()

        # 需要有价格才算数据行
        if not v_price or not isinstance(v_price, (int, float)):
            continue

        price = float(v_price)
        if price <= 0:
            continue

        # 找适用的尺码头
        applicable_sizes = initial_sizes
        for sr in reversed(sorted_size_rows):
            if sr < r:
                applicable_sizes = size_header_rows[sr]
                break

        # 读取各尺码数量
        sizes = {}
        for col_idx, size in applicable_sizes:
            qty = ws.cell(r, col_idx).value
            if qty and isinstance(qty, (int, float)) and qty > 0:
                sizes[str(size)] = int(qty)
                all_sizes_used.add(size)

        total_qty = sum(sizes.values())
        if total_qty == 0:
            continue

        # 解析颜色
        color_str = current_color
        color_code = ''
        color_name = color_str
        mc = re.match(r'^(\d+)\s+(.+)$', color_str)
        if mc:
            color_code = mc.group(1)
            color_name = mc.group(2)

        # 行金额
        line_amount = float(v_amt) if v_amt and isinstance(v_amt, (int, float)) else price * total_qty

        # 分组key：同款+同色+同价 合并，不同价保留独立
        key = (current_item_no, color_str, price)
        if key not in groups:
            groups[key] = {
                'style_code': current_item_no,
                'style': current_code,
                'color_code': color_code,
                'color_name': color_name,
                'color_full': color_str,
                'description': current_desc or f"{current_item_no} {color_name}".strip(),
                'price': price,
                'sizes': {},
                '_line_amount': 0,
            }
        # 合并尺码和金额
        for s, q in sizes.items():
            groups[key]['sizes'][s] = groups[key]['sizes'].get(s, 0) + q
        groups[key]['_line_amount'] += line_amount

    # 构建 items
    items = []
    for g in groups.values():
        g['pieces'] = sum(g['sizes'].values())
        items.append(g)

    # 提取条款和银行信息
    terms = {}
    bank_info = {}
    for r in range(max(data_start, 30), ws.max_row + 1):
        val = str(ws.cell(r, 1).value or '').strip()
        if not val:
            continue
        val_lower = val.lower()
        if 'brand' in val_lower:
            terms['brand'] = val
        elif 'shipment' in val_lower or 'latest' in val_lower:
            terms['shipment'] = val
            m = re.search(r'(\d{1,2}[-/]\w+[-/]\d{4}|\d{4}[-/]\d{1,2}[-/]\d{1,2})', val)
            if m:
                terms['shipment_date'] = m.group(1)
        elif 'port' in val_lower:
            terms['port'] = val
            m = re.search(r'loading:\s*(.+?);', val)
            if m:
                terms['port_loading'] = m.group(1).strip()
            m = re.search(r'destination:\s*(.+)', val)
            if m:
                terms['port_destination'] = m.group(1).strip()
        elif 'delivery' in val_lower or 'fob' in val_lower or 'cif' in val_lower:
            terms['delivery'] = val
        elif 'payment' in val_lower:
            terms['payment'] = val
        elif 'bank name' in val_lower:
            bank_info['bank_name'] = val.split(':', 1)[-1].strip()
        elif 'bank address' in val_lower:
            bank_info['bank_address'] = val.split(':', 1)[-1].strip()
        elif 'beneficiary name' in val_lower:
            bank_info['beneficiary_name'] = val.split(':', 1)[-1].strip()
        elif 'beneficiary address' in val_lower:
            bank_info['beneficiary_address'] = val.split(':', 1)[-1].strip()
        elif 'swift' in val_lower:
            m = re.search(r'SWIFT\s*(?:CODE)?:?\s*([\w]+)', val, re.IGNORECASE)
            if m:
                bank_info['swift'] = m.group(1).strip()
            m = re.search(r'ACCOUNT\s*(?:NO\.?)?:?\s*([\w]+)', val, re.IGNORECASE)
            if m:
                bank_info['account'] = m.group(1).strip()
        elif 'say total' in val_lower:
            terms['amount_in_words'] = val

    total_pieces = sum(item['pieces'] for item in items)
    total_amount = sum(item['pieces'] * item['price'] for item in items)

    return {
        'order_no': order_no,
        'invoice_no': invoice_no,
        'date': date_str,
        'customer': customer_name,
        'customer_address': customer_address,
        'customer_vat': customer_vat,
        'items': items,
        'total_pieces': total_pieces,
        'total_amount': total_amount,
        'terms': terms,
        'bank_info': bank_info,
        'size_headers': sorted(all_sizes_used),
    }


def parse_pi_pdf(filepath):
    """用pdfplumber提取PDF文字，再用AI解析结构"""
    text_content = ""
    with pdfplumber.open(filepath) as pdf:
        for i, page in enumerate(pdf.pages):
            text_content += f"\n--- Page {i+1} ---\n"
            text_content += page.extract_text() or ""
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    text_content += "\t".join(str(c or '') for c in row) + "\n"

    prompt = f"""This is a shoe industry Proforma Invoice (PI). The PI is issued by EMILY (the seller) to a BUYER (customer).
The buyer/customer name appears after "Messers:" — this is the customer placing the order, NOT the seller.
The seller is typically EMILY or BOLLY — do NOT put the seller as the customer.

Extract ALL data and return pure JSON:
{{
  "order_no": "order number (after Order No.)",
  "invoice_no": "invoice number (e.g. MYL-2646)",
  "date": "date string",
  "customer": "BUYER name (after Messers:, NOT the seller)",
  "customer_address": "buyer's address lines",
  "customer_vat": "buyer's VAT/CIF number",
  "items": [
    {{
      "style_code": "style code number",
      "style": "style/model number",
      "color_code": "color number code (e.g. 68)",
      "color_name": "color name (e.g. Marino)",
      "color_full": "full color string (e.g. 68 Marino)",
      "description": "product description",
      "price": unit_price_number,
      "pieces": total_qty_number,
      "sizes": {{"19": 40, "20": 65}}
    }}
  ],
  "total_pieces": total_pairs,
  "total_amount": total_amount,
  "terms": {{
    "brand": "brand line",
    "shipment": "shipment date line",
    "shipment_date": "date only (e.g. 1-Jun-2026)",
    "port": "port of loading/destination line",
    "port_loading": "loading port",
    "port_destination": "destination port",
    "delivery": "delivery terms (FOB/CIF etc)",
    "payment": "payment terms",
    "amount_in_words": "SAY TOTAL line"
  }},
  "bank_info": {{
    "bank_name": "bank name",
    "bank_address": "bank address",
    "beneficiary_name": "beneficiary",
    "beneficiary_address": "beneficiary address",
    "swift": "SWIFT code",
    "account": "account number"
  }},
  "size_headers": [19, 20, 21]
}}

IMPORTANT:
- SPLIT PRICING: If the same style+color has DIFFERENT unit prices for different size ranges (e.g. sizes 19-25 at $9.30 and sizes 26-38 at $9.80), keep them as SEPARATE items. Do NOT merge.
- Split color into color_code (number) and color_name (text). e.g. "68 Marino" → color_code: "68", color_name: "Marino"

PI content:
{text_content[:12000]}

Return ONLY valid JSON, no other text. If certain fields are not found in the PI, use null."""

    result = call_ai(prompt)
    return _parse_ai_json(result)


def parse_pi_image(image_paths):
    """用AI视觉识别PI图片，解析订单数据"""
    if ANTHROPIC_API_KEY:
        images = []
        for path in image_paths:
            ext = os.path.splitext(path)[1].lower()
            media_map = {'.jpg': 'image/jpeg', '.jpeg': 'image/jpeg', '.png': 'image/png', '.webp': 'image/webp', '.heic': 'image/heic'}
            media_type = media_map.get(ext, 'image/jpeg')
            with open(path, 'rb') as f:
                data = base64.b64encode(f.read()).decode('utf-8')
            images.append({'media_type': media_type, 'data': data})
    else:
        images = [{'path': p} for p in image_paths]

    prompt = """These images are a shoe industry Proforma Invoice (PI) or Purchase Order.
The PI is issued by EMILY (the seller) to a BUYER (customer).

Extract ALL data from the images and return pure JSON:
{
  "order_no": "order/PO number",
  "invoice_no": "invoice/reference number",
  "date": "date string",
  "customer": "BUYER name (NOT the seller EMILY)",
  "customer_address": "buyer's address",
  "customer_vat": "buyer's VAT number",
  "currency": "USD or EUR",
  "items": [
    {
      "style_code": "style/article code",
      "style": "model number",
      "color_code": "color number code",
      "color_name": "color name (e.g. pink, marine)",
      "color_full": "full color string",
      "description": "product description",
      "price": unit_price_number,
      "pieces": total_qty_number,
      "sizes": {"19": 40, "20": 65}
    }
  ],
  "total_pieces": total_pairs,
  "total_amount": total_amount,
  "terms": {
    "brand": "brand line",
    "shipment_date": "shipment date",
    "delivery": "delivery terms (FOB/CIF)",
    "payment": "payment terms"
  },
  "size_headers": [19, 20, 21, ...]
}

IMPORTANT:
- Extract EVERY item visible in the images
- SPLIT PRICING: If the same style+color has DIFFERENT unit prices for different size ranges, keep them as SEPARATE items. Do NOT merge rows with different prices.
- Split color into color_code (number) and color_name (text) when possible
- size_headers = sorted list of ALL unique sizes
- Return ONLY valid JSON"""

    result = call_ai(prompt, images=images, timeout=300)
    return _parse_ai_json(result)


# ============================================================
# PO Excel 生成
# ============================================================
def _bank_val(bank_dict, key, default):
    """Get bank value with proper null fallback"""
    v = bank_dict.get(key)
    return v if v else default


def _format_po_no(pi_data):
    """生成PO编号：如果已有PO号直接用，否则从order_no/invoice_no生成"""
    order_no = pi_data.get('order_no', '')
    invoice_no = pi_data.get('invoice_no', '')
    # 如果 order_no 已经是 PO 格式（如 PO100143），直接用
    if order_no and re.match(r'^PO\d+', order_no, re.IGNORECASE):
        return order_no
    # 用 invoice_no 或 order_no 加 PO- 前缀
    ref = invoice_no or order_no
    if ref:
        return f"PO-{ref}" if not ref.upper().startswith('PO') else ref
    return ''


def _apply_notes_to_workbook(wb, output_path, notes, file_type='Excel'):
    """Use AI to apply user notes to a generated workbook.
    Notes = one-time adjustments for this output only, not system-level changes."""
    if not notes or not notes.strip():
        wb.save(output_path)
        return
    # Save temp version first
    wb.save(output_path)
    # Build a summary of current workbook structure for AI
    summary_parts = []
    for ws in wb.worksheets:
        summary_parts.append(f"Sheet: {ws.title}, rows: {ws.max_row}, cols: {ws.max_column}")
        # Sample first 5 rows
        for r in range(1, min(6, ws.max_row + 1)):
            vals = []
            for c in range(1, min(ws.max_column + 1, 20)):
                v = ws.cell(r, c).value
                if v is not None:
                    vals.append(f"{openpyxl.utils.get_column_letter(c)}{r}={v}")
            if vals:
                summary_parts.append("  " + " | ".join(vals))
    structure = "\n".join(summary_parts)

    prompt = f"""You are modifying a generated {file_type} file based on user notes.

Current file structure:
{structure}

User notes: {notes}

Generate Python code using openpyxl to modify the workbook `wb` (already loaded).
Only output the Python code block, no explanation.
Rules:
- Only make changes the user explicitly requested
- Do NOT delete existing data
- Do NOT add malicious content
- Do NOT import os, subprocess, or any system modules
- Only use openpyxl operations (cell values, styles, insert/delete rows/cols, merge)
- Variable `wb` is the workbook, `ws` is wb.active
- If the request is unclear or dangerous, output: # SKIP

Output format:
```python
ws = wb.active
# your code here
```"""

    try:
        result = call_ai(prompt, timeout=30)
        # Extract code block
        code_match = re.search(r'```python\s*\n(.*?)```', result, re.DOTALL)
        if not code_match:
            return
        code = code_match.group(1).strip()
        if '# SKIP' in code:
            return
        # Safety check - block dangerous imports/calls
        dangerous = ['import os', 'import sys', 'import subprocess', 'exec(', 'eval(',
                      '__import__', 'open(', 'system(', 'popen(', 'remove(', 'unlink(']
        if any(d in code for d in dangerous):
            return
        # Execute in restricted namespace
        safe_globals = {
            'wb': wb, 'openpyxl': openpyxl,
            'Font': Font, 'Alignment': Alignment, 'Border': Border,
            'Side': Side, 'PatternFill': PatternFill,
        }
        exec(code, safe_globals)
        wb.save(output_path)
    except Exception:
        # If AI modification fails, keep original file
        pass


def generate_po_excel(pi_data, output_path, pi_format='PR', notes=''):
    """Generate PO Excel — all in English, complete information. pi_format: customer format code (PR/BIS/MYL/custom)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Purchase Order'

    # Styles
    title_font = Font(name='Arial', size=14, bold=True)
    normal_font = Font(name='Arial', size=9)
    small_font = Font(name='Arial', size=8)
    bold_font = Font(name='Arial', size=9, bold=True)
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    light_fill = PatternFill(start_color='F2F4F8', end_color='F2F4F8', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')

    # Column widths
    widths = {'A': 5, 'B': 10, 'C': 10, 'D': 8, 'E': 14, 'F': 20, 'G': 10, 'H': 8}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    sizes = pi_data.get('size_headers', [])
    size_start_col = 9
    for i in range(len(sizes)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(size_start_col + i)].width = 6

    # Amount column (after sizes)
    amt_col = size_start_col + len(sizes)
    ws.column_dimensions[openpyxl.utils.get_column_letter(amt_col)].width = 12

    row = 1
    last_col = openpyxl.utils.get_column_letter(amt_col) if sizes else 'I'

    # === SELLER HEADER ===
    ws.merge_cells(f'A{row}:{last_col}{row}')
    cell = ws.cell(row, 1, EMILY_INFO['company_en'])
    cell.font = title_font
    cell.alignment = center
    row += 1

    ws.merge_cells(f'A{row}:{last_col}{row}')
    ws.cell(row, 1, EMILY_INFO['address']).font = small_font
    ws.cell(row, 1).alignment = center
    row += 2

    # === TITLE ===
    ws.merge_cells(f'A{row}:{last_col}{row}')
    cell = ws.cell(row, 1, 'PURCHASE ORDER')
    cell.font = Font(name='Arial', size=16, bold=True)
    cell.alignment = center
    row += 2

    # === ORDER INFO ===
    terms = pi_data.get('terms') or {}
    info_items = [
        ('PO No.:', _format_po_no(pi_data),
         'Date:', pi_data.get('date', '') or datetime.now().strftime('%d-%b-%Y')),
        ('PI/Ref No.:', pi_data.get('invoice_no', '') or pi_data.get('order_no', ''),
         'Shipment Date:', terms.get('shipment_date') or ''),
    ]

    for left_label, left_val, right_label, right_val in info_items:
        ws.cell(row, 1, left_label).font = bold_font
        ws.cell(row, 1).alignment = right_align
        ws.merge_cells(f'B{row}:C{row}')
        ws.cell(row, 2, left_val).font = normal_font
        ws.cell(row, 5, right_label).font = bold_font
        ws.cell(row, 5).alignment = right_align
        ws.merge_cells(f'F{row}:G{row}')
        ws.cell(row, 6, right_val).font = normal_font
        row += 1

    row += 1

    # === BUYER / SELLER INFO ===
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row, 1, 'BUYER:').font = bold_font
    ws.cell(row, 1).fill = light_fill
    ws.merge_cells(f'E{row}:{last_col}{row}')
    ws.cell(row, 5, 'SELLER:').font = bold_font
    ws.cell(row, 5).fill = light_fill
    row += 1

    buyer_name = pi_data.get('customer') or ''
    buyer_addr = pi_data.get('customer_address') or ''
    buyer_vat = pi_data.get('customer_vat') or ''

    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row, 1, buyer_name).font = bold_font
    ws.merge_cells(f'E{row}:{last_col}{row}')
    ws.cell(row, 5, EMILY_INFO['company_en']).font = bold_font
    row += 1

    if buyer_addr:
        ws.merge_cells(f'A{row}:D{row}')
        ws.cell(row, 1, buyer_addr).font = small_font
        ws.cell(row, 1).alignment = left
    ws.merge_cells(f'E{row}:{last_col}{row}')
    ws.cell(row, 5, EMILY_INFO['address']).font = small_font
    ws.cell(row, 5).alignment = left
    row += 1

    if buyer_vat:
        ws.merge_cells(f'A{row}:D{row}')
        ws.cell(row, 1, f'VAT: {buyer_vat}').font = small_font
        row += 1

    row += 1

    # === PRODUCT TABLE ===
    currency = pi_data.get('currency') or 'USD'
    headers = ['No.', 'Style Code', 'Style', 'Color Code', 'Color', 'Description', f'Unit Price\n({currency})', 'Qty']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row, i, h)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    for i, s in enumerate(sizes):
        col = size_start_col + i
        cell = ws.cell(row, col, s)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    # Amount header
    cell = ws.cell(row, amt_col, f'Amount\n({currency})')
    cell.font = bold_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = center
    row += 1

    # Product data
    for idx, item in enumerate(pi_data.get('items', []), 1):
        ws.cell(row, 1, idx).font = normal_font
        ws.cell(row, 1).border = border
        ws.cell(row, 1).alignment = center

        ws.cell(row, 2, item.get('style_code', '')).font = normal_font
        ws.cell(row, 2).border = border
        ws.cell(row, 2).alignment = center

        ws.cell(row, 3, item.get('style', '')).font = normal_font
        ws.cell(row, 3).border = border
        ws.cell(row, 3).alignment = center

        ws.cell(row, 4, item.get('color_code', '')).font = normal_font
        ws.cell(row, 4).border = border
        ws.cell(row, 4).alignment = center

        ws.cell(row, 5, item.get('color_name', '')).font = normal_font
        ws.cell(row, 5).border = border
        ws.cell(row, 5).alignment = left

        ws.cell(row, 6, item.get('description', '')).font = normal_font
        ws.cell(row, 6).border = border
        ws.cell(row, 6).alignment = left

        price = item.get('price') or 0
        pieces = item.get('pieces') or 0

        ws.cell(row, 7, price).font = normal_font
        ws.cell(row, 7).border = border
        ws.cell(row, 7).alignment = center
        ws.cell(row, 7).number_format = '#,##0.00'

        ws.cell(row, 8, pieces).font = normal_font
        ws.cell(row, 8).border = border
        ws.cell(row, 8).alignment = center

        for i, s in enumerate(sizes):
            col = size_start_col + i
            qty = item.get('sizes', {}).get(str(s), '')
            cell = ws.cell(row, col, qty if qty else '')
            cell.font = small_font
            cell.border = border
            cell.alignment = center

        # Line amount (use pre-calculated if available, otherwise price*qty)
        line_amt = item.get('_line_amount') or (float(price) * int(pieces))
        ws.cell(row, amt_col, line_amt).font = normal_font
        ws.cell(row, amt_col).border = border
        ws.cell(row, amt_col).alignment = center
        ws.cell(row, amt_col).number_format = '#,##0.00'
        row += 1

    # === TOTAL ROW ===
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row, 1, 'TOTAL').font = Font(name='Arial', size=10, bold=True)
    ws.cell(row, 1).alignment = right_align
    ws.cell(row, 1).border = border
    for c in range(2, 8):
        ws.cell(row, c).border = border
    ws.cell(row, 7, '').font = Font(name='Arial', size=10, bold=True)
    ws.cell(row, 7).border = border
    ws.cell(row, 8, pi_data.get('total_pieces') or 0).font = Font(name='Arial', size=10, bold=True)
    ws.cell(row, 8).border = border
    ws.cell(row, 8).alignment = center
    # Size totals
    for i, s in enumerate(sizes):
        col = size_start_col + i
        ws.cell(row, col, '').border = border
    ws.cell(row, amt_col, pi_data.get('total_amount') or 0).font = Font(name='Arial', size=10, bold=True)
    ws.cell(row, amt_col).border = border
    ws.cell(row, amt_col).alignment = center
    ws.cell(row, amt_col).number_format = '#,##0.00'
    row += 1

    # Amount in words
    amount_words = terms.get('amount_in_words') or ''
    if amount_words:
        ws.merge_cells(f'A{row}:{last_col}{row}')
        ws.cell(row, 1, amount_words).font = Font(name='Arial', size=8, italic=True)
        row += 1

    row += 1

    # === TERMS & CONDITIONS ===
    ws.merge_cells(f'A{row}:{last_col}{row}')
    ws.cell(row, 1, 'TERMS & CONDITIONS').font = Font(name='Arial', size=10, bold=True, underline='single')
    row += 1

    terms_display = []
    if terms.get('brand'):
        terms_display.append(terms['brand'])
    if terms.get('shipment'):
        terms_display.append(terms['shipment'])
    if terms.get('port'):
        terms_display.append(terms['port'])
    if terms.get('delivery'):
        terms_display.append(terms['delivery'])
    if terms.get('payment'):
        terms_display.append(terms['payment'])

    if not terms_display:
        terms_display.append('As per PI agreement')

    for t in terms_display:
        ws.merge_cells(f'A{row}:{last_col}{row}')
        ws.cell(row, 1, t).font = normal_font
        ws.cell(row, 1).alignment = left
        row += 1

    row += 1

    # === BANK INFORMATION ===
    ws.merge_cells(f'A{row}:{last_col}{row}')
    ws.cell(row, 1, 'BANK INFORMATION').font = Font(name='Arial', size=10, bold=True, underline='single')
    row += 1

    # 始终使用 EMILY 的收款银行信息（PI中提取的可能是客户的银行）
    bank_lines = [
        ('Bank Name:', EMILY_INFO['bank_name']),
        ('Bank Address:', EMILY_INFO['bank_address']),
        ('Beneficiary Name:', EMILY_INFO['beneficiary']),
        ('Beneficiary Address:', EMILY_INFO['beneficiary_address']),
        ('SWIFT Code:', EMILY_INFO['swift']),
        ('Account No.:', EMILY_INFO['account']),
    ]

    for label, val in bank_lines:
        ws.cell(row, 1, label).font = bold_font
        ws.cell(row, 1).alignment = right_align
        ws.merge_cells(f'B{row}:{last_col}{row}')
        ws.cell(row, 2, val).font = normal_font
        row += 1

    row += 2

    # === SIGNATURES ===
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row, 1, 'For and on behalf of BUYER:').font = bold_font
    ws.merge_cells(f'E{row}:{last_col}{row}')
    ws.cell(row, 5, 'For and on behalf of SELLER:').font = bold_font
    row += 1

    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row, 1, buyer_name).font = normal_font
    ws.merge_cells(f'E{row}:{last_col}{row}')
    ws.cell(row, 5, EMILY_INFO['company_en']).font = normal_font
    row += 3

    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row, 1, 'Signature: ________________________').font = normal_font
    ws.merge_cells(f'E{row}:{last_col}{row}')
    ws.cell(row, 5, 'Signature: ________________________').font = normal_font
    row += 1

    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row, 1, 'Date: ____________________________').font = normal_font
    ws.merge_cells(f'E{row}:{last_col}{row}')
    ws.cell(row, 5, 'Date: ____________________________').font = normal_font

    # Print setup
    try:
        from openpyxl.worksheet.properties import PageSetupProperties
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.orientation = 'landscape'
    except (ImportError, AttributeError):
        pass

    _apply_notes_to_workbook(wb, output_path, notes, 'PI Excel')
    return output_path


# ============================================================
# PO → Packing List
# ============================================================
def generate_packing_list_excel(pi_data, output_path, fty_order='', notes=''):
    """Generate MYL-style Packing List Excel from parsed PO data.
    Each item (style+color) gets a block of rows with box distribution.
    NW/GW left blank for manual fill."""
    wb = openpyxl.Workbook()
    ws = wb.active

    order_no = pi_data.get('invoice_no') or pi_data.get('order_no') or 'ORDER'
    ws.title = str(order_no)[:31]

    # Styles
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    small_font = Font(name='宋体', size=10, bold=True)
    title_font = Font(name='Arial', size=12, bold=True)
    thin = Side(style='thin')
    medium = Side(style='medium')
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # All possible sizes (19-38)
    all_sizes = list(range(19, 39))
    # Determine which sizes are actually used
    used_sizes = set()
    for item in pi_data.get('items', []):
        for s, q in item.get('sizes', {}).items():
            if q and int(q) > 0:
                try:
                    used_sizes.add(int(s))
                except (ValueError, TypeError):
                    pass
    if not used_sizes:
        used_sizes = set(all_sizes)
    size_list = sorted(used_sizes)

    # Size columns: F(col6) onwards
    size_col_start = 6  # F
    size_col_map = {}
    for i, sz in enumerate(all_sizes):
        col_idx = size_col_start + i
        size_col_map[sz] = col_idx

    # Fixed column layout (matches MYL template)
    # A=FTY ORDER, B=CUST PO, C=FTY NO, D=CUST NO, E=COLOR
    # F-Y = sizes 19-38
    # Z=PRS, AA=外箱号, AB=TOTAL CTNS, AC=PER COLOR CTNS, AD=TOTAL PRS, AE=PER COLOR PRS
    # AF=NW, AG=GW, AH-AL=外箱规格, AM=CBM, AN=PER COLOR CBM
    # AO=TNW, AP=TGW, AQ=PER COLOR TGW
    Z = 26   # PRS
    AA = 27  # 外箱号
    AB = 28  # TOTAL CTNS
    AC = 29  # PER COLOR (TOTAL CTNS)
    AD = 30  # TOTAL PRS
    AE = 31  # PER COLOR (TOTAL PRS)
    AF = 32  # NW
    AG = 33  # GW
    AH = 34  # 外箱规格 L
    AI = 35  # *
    AJ = 36  # W
    AK = 37  # *
    AL = 38  # H
    AM = 39  # CBM
    AN = 40  # PER COLOR CBM
    AO = 41  # TNW
    AP = 42  # TGW
    AQ = 43  # PER COLOR TGW

    # Column widths
    col_widths = {'A': 6, 'B': 8, 'C': 7.5, 'D': 10, 'E': 10}
    for sz in all_sizes:
        col_letter = openpyxl.utils.get_column_letter(size_col_map[sz])
        col_widths[col_letter] = 4 if sz in used_sizes else 3
    col_widths.update({
        openpyxl.utils.get_column_letter(Z): 6,
        openpyxl.utils.get_column_letter(AA): 14,
        openpyxl.utils.get_column_letter(AB): 6.5,
        openpyxl.utils.get_column_letter(AC): 6,
        openpyxl.utils.get_column_letter(AD): 11.5,
        openpyxl.utils.get_column_letter(AE): 8,
        openpyxl.utils.get_column_letter(AF): 9,
        openpyxl.utils.get_column_letter(AG): 13,
        openpyxl.utils.get_column_letter(AH): 3.5,
        openpyxl.utils.get_column_letter(AI): 2,
        openpyxl.utils.get_column_letter(AJ): 3.5,
        openpyxl.utils.get_column_letter(AK): 2,
        openpyxl.utils.get_column_letter(AL): 3.5,
        openpyxl.utils.get_column_letter(AM): 8,
        openpyxl.utils.get_column_letter(AN): 8,
        openpyxl.utils.get_column_letter(AO): 9.5,
        openpyxl.utils.get_column_letter(AP): 10,
        openpyxl.utils.get_column_letter(AQ): 13,
    })
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # === Row 1: Title ===
    ws.merge_cells(f'A1:{openpyxl.utils.get_column_letter(AQ)}1')
    ws.cell(1, 1, f'MYL-PACKING LIST-ORDER {order_no}').font = title_font
    ws.row_dimensions[1].height = 26

    # === Row 2: Headers ===
    row2_headers = {
        1: ('FTY ORDER', normal_font),
        2: ('客人PO号 CUST PO', normal_font),
        3: ('型体号 FTY NO.', normal_font),
        4: ('客人型体号 CUST NO.', normal_font),
        5: ('颜色 COLOR', normal_font),
        Z: ('双数 PRS', normal_font),
        AA: ('外箱号', Font(name='Arial', size=12)),
        AB: ('总箱数 TOTAL CTNS', normal_font),
        AC: ('PER COLOR (TOTAL CTNS)', Font(name='Arial', size=8)),
        AD: ('总双数 TOTAL PRS', normal_font),
        AE: ('PER COLOR (TOTAL PRS)', normal_font),
        AF: ('净重 NW(KGS)', Font(name='Arial', size=11)),
        AG: ('毛重 GW(KGS)', Font(name='Arial', size=11)),
        AH: ('外箱规格 CM', normal_font),
        AM: ('立方数\n（CBM)', normal_font),
        AN: ('PER COLOR (TOTAL CBM)', normal_font),
        AO: ('总净重 TNW(KGS)', Font(name='Arial', size=9)),
        AP: ('总毛重 TGW(KGS)', Font(name='Arial', size=9)),
        AQ: ('每色总毛重 TGW(KGS)', Font(name='Arial', size=9)),
    }
    for col_idx, (text, font) in row2_headers.items():
        c = ws.cell(2, col_idx, text)
        c.font = font
        c.alignment = center
        c.border = border_thin
    # Merge AH-AL for 外箱规格
    ws.merge_cells(f'{openpyxl.utils.get_column_letter(AH)}2:{openpyxl.utils.get_column_letter(AL)}2')
    ws.row_dimensions[2].height = 29

    # === Row 3: Size numbers ===
    for sz in all_sizes:
        col = size_col_map[sz]
        c = ws.cell(3, col, sz)
        c.font = small_font
        c.alignment = center
        c.border = border_thin
    ws.row_dimensions[3].height = 29

    # === Data rows ===
    row = 4
    data_start_row = 4
    items = pi_data.get('items', [])

    # Group items by style
    # Keep order but track style groups for FTY ORDER display
    prev_style = None

    for item_idx, item in enumerate(items):
        style_code = item.get('style_code', '') or ''
        color_code = item.get('color_code', '') or ''
        color_name = item.get('color_name', '') or ''
        color_display = f'{color_code} {color_name}'.strip() if color_code or color_name else ''
        description = item.get('description', '') or item.get('style', '') or ''

        # Build fty_no from style_code (remove dots for FTY NO)
        cust_no = style_code  # e.g. 41.727
        fty_no = ''  # To be filled manually or derived

        # Get size quantities
        size_qtys = {}
        for s, q in item.get('sizes', {}).items():
            try:
                sz = int(float(s))
                qty = int(float(q)) if q else 0
                if qty > 0:
                    size_qtys[sz] = qty
            except (ValueError, TypeError):
                pass

        if not size_qtys:
            continue

        # === Box distribution algorithm ===
        # Box capacity based on shoe size (matches MYL standard)
        def _box_cap(sz):
            if sz <= 22: return 40
            if sz <= 25: return 30
            if sz <= 27: return 30
            if sz <= 30: return 28
            if sz <= 33: return 18
            return 16  # 34-38

        box_rows = []  # list of (size_dict, num_boxes)
        remainders = {}

        for sz in sorted(size_qtys.keys()):
            qty = size_qtys[sz]
            cap = _box_cap(sz)

            full_boxes = qty // cap
            remainder = qty % cap

            if full_boxes > 0:
                box_rows.append(({sz: cap}, full_boxes))
            if remainder > 0:
                remainders[sz] = remainder

        # Combine remainders into mixed boxes
        if remainders:
            rem_sizes = sorted(remainders.keys())
            current_box = {}
            current_total = 0
            cap = _box_cap(rem_sizes[0]) if rem_sizes else 40  # use first remainder size's capacity

            for sz in rem_sizes:
                qty = remainders[sz]
                if current_total + qty <= cap:
                    current_box[sz] = qty
                    current_total += qty
                else:
                    if current_box:
                        box_rows.append((current_box, 1))
                    current_box = {sz: qty}
                    current_total = qty
            if current_box:
                box_rows.append((current_box, 1))

        # Write color block
        block_start = row
        first_row = True

        for br_idx, (sizes_in_box, num_boxes) in enumerate(box_rows):
            if first_row:
                # FTY ORDER (only on first item of each style group or new FTY ORDER)
                if fty_order:
                    ws.cell(row, 1, fty_order).font = normal_font
                    ws.cell(row, 1).border = Border(left=medium, right=thin, top=thin, bottom=thin)
                # CUST PO
                po_no = pi_data.get('invoice_no') or pi_data.get('order_no') or ''
                po_clean = re.sub(r'^(PO|PV|ORDER)\s*', '', str(po_no), flags=re.IGNORECASE).strip()
                # Try to convert to int for cleaner display
                try:
                    po_clean = int(po_clean)
                except (ValueError, TypeError):
                    pass
                ws.cell(row, 2, po_clean).font = normal_font
                ws.cell(row, 2).alignment = center
                ws.cell(row, 2).border = border_thin
                # FTY NO
                ws.cell(row, 3, fty_no).font = normal_font
                ws.cell(row, 3).alignment = center
                ws.cell(row, 3).border = border_thin
                # CUST NO (try to keep as float like 41.727)
                try:
                    cust_no_val = float(cust_no) if cust_no else ''
                except (ValueError, TypeError):
                    cust_no_val = cust_no
                ws.cell(row, 4, cust_no_val).font = normal_font
                ws.cell(row, 4).alignment = center
                ws.cell(row, 4).border = border_thin
                # COLOR
                ws.cell(row, 5, color_display).font = normal_font
                ws.cell(row, 5).alignment = center
                ws.cell(row, 5).border = border_thin
                first_row = False

            # Size quantities
            for sz, qty in sizes_in_box.items():
                col = size_col_map.get(sz)
                if col:
                    ws.cell(row, col, qty).font = normal_font
                    ws.cell(row, col).alignment = center
                    ws.cell(row, col).border = border_thin

            # Z: PRS = SUM of sizes
            z_letter = openpyxl.utils.get_column_letter(Z)
            f_letter = openpyxl.utils.get_column_letter(size_col_start)
            y_letter = openpyxl.utils.get_column_letter(size_col_start + len(all_sizes) - 1)
            ws.cell(row, Z, f'=SUM({f_letter}{row}:{y_letter}{row})').font = normal_font
            ws.cell(row, Z).alignment = center
            ws.cell(row, Z).border = border_thin

            # AB: boxes count
            ab_letter = openpyxl.utils.get_column_letter(AB)
            ws.cell(row, AB, num_boxes).font = normal_font
            ws.cell(row, AB).alignment = center
            ws.cell(row, AB).border = border_thin

            # AD: total PRS = Z * AB
            ad_letter = openpyxl.utils.get_column_letter(AD)
            ws.cell(row, AD, f'={z_letter}{row}*{ab_letter}{row}').font = normal_font
            ws.cell(row, AD).alignment = center
            ws.cell(row, AD).border = border_thin

            # AH-AL: box dimensions (default 60*40*30, blank for manual fill)
            ah_letter = openpyxl.utils.get_column_letter(AH)
            aj_letter = openpyxl.utils.get_column_letter(AJ)
            al_letter = openpyxl.utils.get_column_letter(AL)
            ai_letter = openpyxl.utils.get_column_letter(AI)
            ak_letter = openpyxl.utils.get_column_letter(AK)
            ws.cell(row, AH, 60).font = normal_font
            ws.cell(row, AH).alignment = center
            ws.cell(row, AI, '*').font = normal_font
            ws.cell(row, AI).alignment = center
            ws.cell(row, AJ, 40).font = normal_font
            ws.cell(row, AJ).alignment = center
            ws.cell(row, AK, '*').font = normal_font
            ws.cell(row, AK).alignment = center
            ws.cell(row, AL, 30).font = normal_font
            ws.cell(row, AL).alignment = center

            # AM: CBM = L*W*H/1000000 * boxes
            am_letter = openpyxl.utils.get_column_letter(AM)
            ws.cell(row, AM, f'={ah_letter}{row}*{aj_letter}{row}*{al_letter}{row}/1000000*{ab_letter}{row}').font = normal_font
            ws.cell(row, AM).alignment = center
            ws.cell(row, AM).border = border_thin

            # AO: TNW = boxes * NW
            af_letter = openpyxl.utils.get_column_letter(AF)
            ag_letter = openpyxl.utils.get_column_letter(AG)
            ao_letter = openpyxl.utils.get_column_letter(AO)
            ap_letter = openpyxl.utils.get_column_letter(AP)
            ws.cell(row, AO, f'={ab_letter}{row}*{af_letter}{row}').font = normal_font
            ws.cell(row, AO).alignment = center
            ws.cell(row, AO).border = border_thin
            # AP: TGW = boxes * GW
            ws.cell(row, AP, f'={ab_letter}{row}*{ag_letter}{row}').font = normal_font
            ws.cell(row, AP).alignment = center
            ws.cell(row, AP).border = border_thin

            row += 1

        block_end = row - 1

        # Add PER COLOR formulas on the first row of the block
        if block_start <= block_end:
            # AC: PER COLOR TOTAL CTNS
            ac_letter = openpyxl.utils.get_column_letter(AC)
            ws.cell(block_start, AC, f'=SUM({ab_letter}{block_start}:{ab_letter}{block_end})').font = normal_font
            ws.cell(block_start, AC).alignment = center
            ws.cell(block_start, AC).border = border_thin

            # AE: PER COLOR TOTAL PRS
            ae_letter = openpyxl.utils.get_column_letter(AE)
            ws.cell(block_start, AE, f'=SUM({ad_letter}{block_start}:{ad_letter}{block_end})').font = normal_font
            ws.cell(block_start, AE).alignment = center
            ws.cell(block_start, AE).border = border_thin

            # AN: PER COLOR TOTAL CBM
            an_letter = openpyxl.utils.get_column_letter(AN)
            ws.cell(block_start, AN, f'=SUM({am_letter}{block_start}:{am_letter}{block_end})').font = normal_font
            ws.cell(block_start, AN).alignment = center
            ws.cell(block_start, AN).border = border_thin

            # AQ: PER COLOR TGW
            aq_letter = openpyxl.utils.get_column_letter(AQ)
            ws.cell(block_start, AQ, f'=SUM({ap_letter}{block_start}:{ap_letter}{block_end})').font = normal_font
            ws.cell(block_start, AQ).alignment = center
            ws.cell(block_start, AQ).border = border_thin

            # Merge cells for multi-row blocks
            if block_end > block_start:
                for merge_col in [3, 4, 5]:  # C, D, E
                    col_l = openpyxl.utils.get_column_letter(merge_col)
                    ws.merge_cells(f'{col_l}{block_start}:{col_l}{block_end}')
                # Merge PER COLOR columns
                for merge_col in [AC, AE, AN, AQ]:
                    col_l = openpyxl.utils.get_column_letter(merge_col)
                    ws.merge_cells(f'{col_l}{block_start}:{col_l}{block_end}')

    # === Total row ===
    total_row = row
    ab_letter = openpyxl.utils.get_column_letter(AB)
    ac_letter = openpyxl.utils.get_column_letter(AC)
    ad_letter = openpyxl.utils.get_column_letter(AD)
    ae_letter = openpyxl.utils.get_column_letter(AE)
    am_letter = openpyxl.utils.get_column_letter(AM)
    an_letter = openpyxl.utils.get_column_letter(AN)
    ao_letter = openpyxl.utils.get_column_letter(AO)
    ap_letter = openpyxl.utils.get_column_letter(AP)
    aq_letter = openpyxl.utils.get_column_letter(AQ)

    for col_idx, col_letter in [(AB, ab_letter), (AC, ac_letter), (AD, ad_letter),
                                 (AE, ae_letter), (AM, am_letter), (AN, an_letter),
                                 (AO, ao_letter), (AP, ap_letter), (AQ, aq_letter)]:
        ws.cell(total_row, col_idx,
                f'=SUM({col_letter}{data_start_row}:{col_letter}{total_row-1})').font = bold_font
        ws.cell(total_row, col_idx).alignment = center
        ws.cell(total_row, col_idx).border = border_thin

    # Print setup
    try:
        from openpyxl.worksheet.properties import PageSetupProperties
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.orientation = 'landscape'
    except (ImportError, AttributeError):
        pass

    _apply_notes_to_workbook(wb, output_path, notes, 'Packing List')
    return output_path


# ============================================================
# PI → CI (Commercial Invoice)
# ============================================================
def generate_ci_from_pi(pi_path, output_path, notes=''):
    """Convert PI Excel to CI Excel.
    - Change title from PROFORMA INVOICE to COMMERCIAL INVOICE
    - Insert shipping detail rows after header
    - Add bank info and declaration at bottom
    Processes ALL sheets in the workbook.
    """
    import copy as _copy
    wb = openpyxl.load_workbook(pi_path)

    for ws in wb.worksheets:
        _convert_sheet_pi_to_ci(ws, wb)

    _apply_notes_to_workbook(wb, output_path, notes, 'Commercial Invoice')
    return output_path


def _convert_sheet_pi_to_ci(ws, wb):
    """Convert a single PI sheet to CI format."""

    # Step 1: Find and replace title in A1
    a1_val = ws.cell(1, 1).value or ''
    if 'PROFORMA' in a1_val.upper():
        ws.cell(1, 1).value = a1_val.replace('PROFORMA INVOICE', 'COMMERCIAL  INVOICE').replace('Proforma Invoice', 'COMMERCIAL  INVOICE').replace('proforma invoice', 'COMMERCIAL  INVOICE')
    elif 'INVOICE' not in a1_val.upper():
        ws.cell(1, 1).value = a1_val + '\nCOMMERCIAL  INVOICE'
    else:
        ws.cell(1, 1).value = a1_val.replace('INVOICE', 'INVOICE').replace('PROFORMA', 'COMMERCIAL ')

    # Step 2: Find the row with column headers (Style, Color, Description...)
    header_row = None
    for r in range(1, min(ws.max_row + 1, 15)):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and 'Style' in str(v):
                header_row = r
                break
        if header_row:
            break

    if not header_row:
        header_row = 2  # fallback

    # Step 3: Insert 5 rows before header_row for shipping details
    ws.insert_rows(header_row, 5)
    ship_row = header_row

    shipping_fields = [
        ('PORT OF LOADING:  ', 'YANGON MYANMAR', 'EX-FACTORY DATE: '),
        ('PORT OF DISCHARGE:  ', 'ALGECIRAS PORT, SPAIN', 'SHIPPED:  BY SEA'),
        ('SHIPPED BY: ', '', 'ETD: '),
        ('CONTAINER NO: ', '', ''),
        ('DELIVERY TERMS:', 'FOB YANGON PORT, MYANMAR (INCOTERMS 2010)', ''),
    ]
    normal_font = Font(name='Arial', size=10)
    for i, (label, val, right_val) in enumerate(shipping_fields):
        r = ship_row + i
        ws.cell(r, 1, label).font = normal_font
        if val:
            ws.cell(r, 3, val).font = normal_font
        if right_val:
            ws.cell(r, 8, right_val).font = normal_font

    # Extract order number for the last shipping row
    order_no = ''
    a1 = ws.cell(1, 1).value or ''
    import re as _re
    m = _re.search(r'Order\s*No\.?\s*:?\s*([^\s\n]+)', a1, _re.IGNORECASE)
    if m:
        order_no = m.group(1).strip()
    ws.cell(ship_row + 4, 8, f'Order No.: {order_no}').font = normal_font

    # Step 4: Update DATE in header to current date (or leave as is)
    # Find "DATE:" in A1 text and potentially update
    # For now, leave the original date

    # Step 5: Add bank info and declaration at bottom (only if not already present)
    has_bank = False
    has_decl = False
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v:
            vs = str(v)
            if 'Bank information' in vs or 'Bank name' in vs:
                has_bank = True
            if 'exporter' in vs.lower() and 'origin' in vs.lower():
                has_decl = True

    last_data_row = ws.max_row
    # Find "Total Order" row
    total_row = None
    for r in range(1, last_data_row + 1):
        for c in range(1, 8):
            v = ws.cell(r, c).value
            if v and 'Total Order' in str(v):
                total_row = r
    if not total_row:
        total_row = last_data_row

    # Add SAY TOTAL row (after Total Order, before any existing bank info)
    if not has_bank:
        say_row = total_row + 1
        ws.cell(say_row, 1, 'SAY TOTAL: ').font = normal_font

        bank_row = say_row + 1
        bank_lines = [
            '(6) Bank information:',
            '    Bank name: E.SUN COMMERCIAL BANK LTD., HONG KONG BRANCH',
            '     Bank address: SUITE 2805, 28F, TOWER 6, THE GATEWAY, 9 CANTON ROAD, TSIMSHATSUI, KOWLOON, HONG KONG',
            '     Beneficiary Name: EMILY HONG KONG LIMITED',
            '     Beneficiary address: FLAT/RM 20, 8/F, YALE INDUSTRIAL CENTRE, 61-63 AU PUI WAN STREET, FO TAN, SHATIN, NT, HONG KONG ',
            f'     SWIFT CODE: {EMILY_INFO["swift"]}    ACCOUNT NO.: {EMILY_INFO["account"]}',
        ]
        for i, line in enumerate(bank_lines):
            ws.cell(bank_row + i, 1, line).font = normal_font

        if not has_decl:
            decl_row = bank_row + len(bank_lines) + 1
            decl_lines = [
                'The exporter BOLLY (HONG KONG) COMPANY LIMITED  (Number of Registered Exporter (MMREX00071) of the ',
                'products covered by this document declares that, except where otherwise clearly indicated, these products are ',
                'of Myanmar preferential origin according to rules  of origin of the Generalized System of ',
                'Preferences of the  European Union and that the origin criterion met is W6402.',
            ]
            ws.cell(decl_row, 1, '\n'.join(decl_lines)).font = normal_font
            ws.cell(decl_row, 1).alignment = Alignment(wrap_text=True, vertical='top')


# ============================================================
# 手写原材料 → 生产指令单
# ============================================================
def parse_handwritten_materials(image_paths):
    """识别手写原材料采购需求表"""
    prompt = """这些是鞋厂的采购需求表和手写记录。请仔细识别所有印刷和手写内容，返回纯JSON格式：
{
  "order_no": "订单号(如MYL-2646)",
  "date": "日期",
  "total_pairs": 总双数,
  "materials": [
    {
      "category": "分类(面料PU/五金/织带/皮料/其他)",
      "sample_color": "色卡描述",
      "material_code": "材料编号",
      "style_code": "型体号",
      "unit_usage": "单耗用量",
      "percentage": "预补百分比",
      "total_usage": "合计用量",
      "ordered_qty": "订购量(手写数字)",
      "unit": "单位(码/个/套/SF等)"
    }
  ],
  "hardware": [
    {
      "name": "物料名称",
      "spec": "规格",
      "color": "颜色",
      "quantity": 数量,
      "unit": "单位",
      "unit_price": 单价,
      "total_price": 总价,
      "supplier": "供应商"
    }
  ],
  "production_notes": ["备注内容"]
}
仔细识别所有手写数字。最终只输出纯JSON。"""

    if ANTHROPIC_API_KEY:
        images = []
        for path in image_paths:
            with open(path, 'rb') as f:
                b64 = base64.b64encode(f.read()).decode()
            ext = path.lower().rsplit('.', 1)[-1]
            media_type = 'image/jpeg' if ext in ('jpg', 'jpeg') else 'image/png'
            images.append({'media_type': media_type, 'data': b64})
        result = call_ai(prompt, images=images, timeout=180)
    else:
        result = call_ai(prompt, images=[{'path': p} for p in image_paths], timeout=600)

    return _parse_ai_json(result)


def generate_production_sheet_html(material_data, notes=''):
    """生成生产指令单HTML"""
    order_no = material_data.get('order_no', 'MYL-XXXX')
    date = material_data.get('date', datetime.now().strftime('%Y/%m/%d'))
    total_pairs = material_data.get('total_pairs', 0) or 0
    materials = material_data.get('materials', [])
    hardware = material_data.get('hardware', [])

    categories = {}
    for m in materials:
        cat = m.get('category', '其他')
        categories.setdefault(cat, []).append(m)

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
@page {{ size: A4 landscape; margin: 10mm; }}
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: -apple-system, "Microsoft YaHei", "SimHei", Arial, sans-serif; font-size: 12px; padding: 20px; color: #1a1a1a; }}
h1 {{ text-align: center; font-size: 20px; margin: 0 0 8px; letter-spacing: 2px; }}
.company {{ text-align: center; font-size: 11px; color: #666; margin-bottom: 16px; }}
.header {{ display: flex; justify-content: space-between; margin-bottom: 12px; font-size: 13px; padding: 8px 12px; background: #f5f7fa; border-radius: 6px; }}
.header span {{ margin-right: 24px; }}
.header b {{ color: #333; }}
table {{ border-collapse: collapse; width: 100%; margin-bottom: 16px; }}
th, td {{ border: 1px solid #ccc; padding: 5px 8px; text-align: center; font-size: 11px; }}
th {{ background: #e8edf3; font-weight: 600; color: #333; }}
tr:nth-child(even) {{ background: #fafbfc; }}
.section-title {{ font-size: 14px; font-weight: 700; margin: 16px 0 6px; padding-bottom: 4px; border-bottom: 2px solid #4a6fa5; color: #2c3e50; }}
td.left {{ text-align: left; }}
.ordered {{ font-weight: 700; color: #c0392b; font-size: 12px; }}
.notes {{ margin-top: 16px; font-size: 12px; line-height: 1.8; padding: 10px; background: #fffbe6; border: 1px solid #ffe58f; border-radius: 4px; }}
.footer {{ margin-top: 24px; display: flex; justify-content: space-between; font-size: 12px; }}
.footer div {{ border-bottom: 1px solid #999; padding-bottom: 4px; min-width: 200px; }}
@media print {{ body {{ padding: 0; }} .no-print {{ display: none; }} }}
.print-btn {{ position: fixed; top: 16px; right: 16px; padding: 8px 20px; background: #4a6fa5; color: #fff; border: none; border-radius: 6px; cursor: pointer; font-size: 13px; }}
.print-btn:hover {{ background: #3a5f95; }}
</style>
</head>
<body>
<button class="print-btn no-print" onclick="window.print()">Print / Export PDF</button>

<h1>EMILY Production Instruction Sheet</h1>
<div class="company">{EMILY_INFO['company_en']} &nbsp;|&nbsp; {EMILY_INFO['address']}</div>

<div class="header">
    <span><b>Order No.:</b> {order_no}</span>
    <span><b>Date:</b> {date}</span>
    <span><b>Total Qty:</b> {total_pairs:,} prs</span>
</div>

<div class="section-title">I. Raw Material Requirements ({len(materials)} items)</div>
<table>
<tr>
    <th style="width:4%">No.</th>
    <th style="width:10%">Category</th>
    <th style="width:18%">Material Code</th>
    <th style="width:8%">Style</th>
    <th style="width:8%">Usage/pr</th>
    <th style="width:6%">Spare%</th>
    <th style="width:10%">Total Usage</th>
    <th style="width:10%">Order Qty</th>
    <th style="width:6%">Unit</th>
    <th style="width:20%">Color/Note</th>
</tr>"""

    idx = 1
    for cat, cat_items in categories.items():
        for item in cat_items:
            html += f"""
<tr>
    <td>{idx}</td>
    <td>{cat}</td>
    <td class="left">{item.get('material_code', '')}</td>
    <td>{item.get('style_code', '')}</td>
    <td>{item.get('unit_usage', '')}</td>
    <td>{item.get('percentage', '')}</td>
    <td>{item.get('total_usage', '')}</td>
    <td class="ordered">{item.get('ordered_qty', '')}</td>
    <td>{item.get('unit', '')}</td>
    <td class="left">{item.get('sample_color', '')}</td>
</tr>"""
            idx += 1

    html += "</table>"

    if hardware:
        total_hw_cost = sum(float(hw.get('total_price', 0) or 0) for hw in hardware)
        html += f"""
<div class="section-title">II. Hardware & Accessories ({len(hardware)} items, Total ¥{total_hw_cost:,.2f})</div>
<table>
<tr>
    <th style="width:4%">No.</th>
    <th style="width:15%">Item</th>
    <th style="width:12%">Spec</th>
    <th style="width:10%">Color</th>
    <th style="width:8%">Qty</th>
    <th style="width:6%">Unit</th>
    <th style="width:8%">Unit Price</th>
    <th style="width:10%">Total</th>
    <th style="width:15%">Supplier</th>
</tr>"""
        for i, hw in enumerate(hardware, 1):
            html += f"""
<tr>
    <td>{i}</td>
    <td class="left">{hw.get('name', '')}</td>
    <td>{hw.get('spec', '')}</td>
    <td>{hw.get('color', '')}</td>
    <td>{hw.get('quantity', '')}</td>
    <td>{hw.get('unit', '')}</td>
    <td>¥{hw.get('unit_price', '')}</td>
    <td>¥{hw.get('total_price', '')}</td>
    <td class="left">{hw.get('supplier', '')}</td>
</tr>"""
        html += "</table>"

    notes = material_data.get('production_notes', [])
    if notes:
        html += '<div class="notes"><b>Notes:</b><br>'
        for n in notes:
            html += f'&bull; {n}<br>'
        html += '</div>'

    html += f"""
<div class="footer">
    <div>Prepared by:</div>
    <div>Reviewed by:</div>
    <div>Date: {datetime.now().strftime('%Y/%m/%d')}</div>
</div>
</body>
</html>"""
    return html


# ============================================================
# Quotation → COG Overview 转换
# ============================================================
def parse_quotation(filepath):
    """解析 Emily 报价单 Excel，提取所有款式/颜色/尺码段/价格及产品图片"""
    if filepath.lower().endswith('.xls') and not filepath.lower().endswith('.xlsx'):
        filepath = convert_xls_to_xlsx(filepath)

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # 提取图片：按行映射到 style_number
    # 先建 row -> style_number 映射（需要先扫一遍找到header和数据行）
    _img_dir = os.path.join(app.config['UPLOAD_FOLDER'], f'quotation_imgs_{int(time.time())}')
    os.makedirs(_img_dir, exist_ok=True)

    # 读取头部信息
    season = ''
    vendor = ''
    date_str = ''
    for r in range(1, 6):
        for c in range(1, 5):
            v = str(ws.cell(r, c).value or '').strip()
            if not v:
                continue
            m = re.search(r'SEASON\s*[:：]\s*(.+)', v, re.IGNORECASE)
            if m:
                season = m.group(1).strip()
            m = re.search(r'VENDOR\s*(?:NAME)?\s*[:：]\s*(.+)', v, re.IGNORECASE)
            if m:
                vendor = m.group(1).strip()
            m = re.search(r'DATE\s*[:：]\s*(.+)', v, re.IGNORECASE)
            if m:
                date_str = m.group(1).strip()

    # 找列头行
    header_row = 0
    for r in range(1, 10):
        row_text = ' '.join(str(ws.cell(r, c).value or '').lower() for c in range(1, 11))
        if 'style' in row_text and 'price' in row_text:
            header_row = r
            break
    if not header_row:
        return None

    # 自动检测列位置
    col_map = {}
    for c in range(1, ws.max_column + 1):
        cv = str(ws.cell(header_row, c).value or '').lower().strip()
        if not cv:
            continue
        if 'fty' in cv or ('style' in cv and 'no' in cv and 'number' not in cv):
            col_map['fty'] = c
        elif 'style name' in cv or cv == 'style name':
            col_map['name'] = c
        elif 'style number' in cv or cv == 'style number':
            col_map['number'] = c
        elif 'color' in cv:
            col_map['color'] = c
        elif 'size range' in cv:
            col_map['range'] = c
        elif cv == 'size' or cv == 'sizes':
            col_map['size'] = c
        elif 'price' in cv:
            col_map['price'] = c
        elif 'remark' in cv:
            col_map['remark'] = c

    fty_col = col_map.get('fty', 1)
    name_col = col_map.get('name', 3)
    number_col = col_map.get('number', 4)
    color_col = col_map.get('color', 6)
    range_col = col_map.get('range', 7)
    size_col = col_map.get('size', 8)
    price_col = col_map.get('price', 9)
    remark_col = col_map.get('remark', 10)

    # 提取图片：row -> style_number -> 保存第一张图
    row_to_style = {}
    for r in range(header_row + 1, ws.max_row + 1):
        v_name = ws.cell(r, name_col).value
        v_number = ws.cell(r, number_col).value
        if v_name and v_number:
            row_to_style[r] = str(v_number).strip()

    # Extract images: map each image to its row number
    row_image_path = {}  # excel_row -> saved image path
    style_image_path = {}  # style_number -> saved image path (fallback)
    _img_counter = 0
    for img in ws._images:
        try:
            anc = img.anchor
            if not hasattr(anc, '_from'):
                continue
            excel_row = anc._from.row + 1  # 0-indexed to 1-indexed
            # Save image
            from openpyxl.drawing.image import Image as XlImage
            img_data = img._data()
            ext = 'jpg'
            if hasattr(img, 'format') and img.format:
                ext = img.format.lower()
            elif img.path and '.' in img.path:
                ext = img.path.rsplit('.', 1)[-1].lower()
            _img_counter += 1
            img_path = os.path.join(_img_dir, f'row_{excel_row}_{_img_counter}.{ext}')
            with open(img_path, 'wb') as f:
                f.write(img_data)
            row_image_path[excel_row] = img_path
            # Also map to style if this row has a style
            style_no = row_to_style.get(excel_row)
            if style_no and style_no not in style_image_path:
                style_image_path[style_no] = img_path
        except Exception:
            pass

    # 解析数据行，构建条目列表
    entries = []  # 每个entry: {name, number, colors:[], size_range, price_tiers:[{range, price}], remark}
    current_entry = None
    last_size_range = ''  # 记录上一个款式的size_range，用于变体继承

    for r in range(header_row + 1, ws.max_row + 1):
        v_name = ws.cell(r, name_col).value
        v_number = ws.cell(r, number_col).value
        v_color = ws.cell(r, color_col).value
        v_range = ws.cell(r, range_col).value
        v_size = ws.cell(r, size_col).value
        v_price = ws.cell(r, price_col).value
        v_remark = ws.cell(r, remark_col).value

        # 新条目：有 style name + style number
        if v_name and v_number:
            # 保存上一个
            if current_entry and current_entry['price_tiers']:
                entries.append(current_entry)

            name_str = str(v_name).strip()
            # 清理换行和多余空格
            name_str = re.sub(r'\s+', ' ', name_str).strip()

            colors = []
            if v_color:
                color_text = str(v_color).strip()
                for line in color_text.split('\n'):
                    line = line.strip()
                    if line:
                        colors.append(line)

            size_range_str = str(v_range).strip().rstrip('#') if v_range else ''
            # 没有 size_range 时继承同款式上一个变体的范围
            if not size_range_str and last_size_range:
                size_range_str = last_size_range
            if size_range_str:
                last_size_range = size_range_str

            style_no_str = str(v_number).strip() if v_number else ''
            # Build per-color image mapping
            color_images = {}
            # Check if this row or nearby rows have images
            row_img = row_image_path.get(r, '')
            if row_img:
                # Assign this image to all colors in this cell
                for c_name in colors:
                    color_images[c_name] = row_img
            else:
                # Fallback to style-level image
                fallback = style_image_path.get(style_no_str, '')
                if fallback:
                    for c_name in colors:
                        color_images[c_name] = fallback
            current_entry = {
                'name': name_str,
                'number': style_no_str,
                'colors': colors,
                'color_images': color_images,
                'size_range': size_range_str,
                'price_tiers': [],
                'remark': str(v_remark).strip() if v_remark else '',
                'image_path': style_image_path.get(style_no_str, ''),
                'source_row': r,
            }

            # 读取第一个价格段
            if v_size and v_price:
                size_tier = str(v_size).strip().rstrip('#')
                try:
                    current_entry['price_tiers'].append({
                        'range': size_tier,
                        'price': float(v_price),
                    })
                except (ValueError, TypeError):
                    pass
            elif v_price:
                # 只有价格没有单独的size列（如只有一个尺码段）
                try:
                    current_entry['price_tiers'].append({
                        'range': size_range_str,
                        'price': float(v_price),
                    })
                except (ValueError, TypeError):
                    pass

        elif v_size and v_price and current_entry:
            # 追加价格段到当前条目
            size_tier = str(v_size).strip().rstrip('#')
            try:
                current_entry['price_tiers'].append({
                    'range': size_tier,
                    'price': float(v_price),
                })
            except (ValueError, TypeError):
                pass

    # 保存最后一个条目
    if current_entry and current_entry['price_tiers']:
        entries.append(current_entry)

    return {
        'season': season,
        'vendor': vendor,
        'date': date_str,
        'entries': entries,
    }


def _make_thumbnail(src_path, max_w=80, max_h=60):
    """用 Pillow 将图片缩为缩略图，返回临时文件路径"""
    try:
        from PIL import Image as PILImage
        img = PILImage.open(src_path)
        img.thumbnail((max_w, max_h), PILImage.LANCZOS)
        thumb_path = src_path.rsplit('.', 1)[0] + '_thumb.jpg'
        img.convert('RGB').save(thumb_path, 'JPEG', quality=85)
        return thumb_path
    except Exception:
        return None


def generate_cog_excel(quotation_data, output_path, brand_prefix='bisgaard', notes=''):
    """将报价单展开为 COG Overview Excel（每颜色×每尺码=一行，含产品图片）"""
    from openpyxl.drawing.image import Image as XlImage

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ark1'

    # Styles
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    header_font = Font(name='Arial', size=10, bold=True)
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    # Column widths (A列加宽以容纳图片)
    widths = {'A': 18, 'B': 28, 'C': 12, 'D': 18, 'E': 14, 'F': 8, 'G': 10, 'H': 28}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Header info
    ws.cell(2, 1, f"SEASON: {quotation_data.get('season', '')}").font = bold_font
    ws.cell(3, 1, f"VENDOR: {quotation_data.get('vendor', '')}").font = bold_font
    ws.cell(4, 1, f"DATE:{quotation_data.get('date', '')}").font = bold_font

    # Column headers (row 6)
    headers = ['PHOTO', 'STYLE NAME', 'STYLE NO.', 'COLOR CODE', 'SIZE RANGE', 'SIZE', 'COG', 'ADDITIONAL COMMENT']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(6, i, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    row = 7
    THUMB_W = 120  # 缩略图宽度(像素)
    THUMB_H = 85   # 缩略图高度(像素)
    IMG_ROW_HEIGHT = 68  # 图片行行高(点)

    for entry in quotation_data.get('entries', []):
        style_name = entry['name']
        style_no = entry['number']
        size_range = entry['size_range']
        remark = entry.get('remark', '')
        price_tiers = entry['price_tiers']
        colors = entry['colors']
        image_path = entry.get('image_path', '')
        color_images = entry.get('color_images', {})

        # 构建 style name with brand prefix
        full_name = f"{brand_prefix} {style_name}" if brand_prefix else style_name

        # 解析完整尺码范围
        range_match = re.match(r'(\d+)\s*[-–]\s*(\d+)', size_range)
        if not range_match:
            continue
        range_start = int(range_match.group(1))
        range_end = int(range_match.group(2))
        all_sizes = list(range(range_start, range_end + 1))

        # 为每个尺码确定价格（根据price_tiers）
        size_price_map = {}
        for tier in price_tiers:
            tier_match = re.match(r'(\d+)\s*[-–]\s*(\d+)', tier['range'])
            if tier_match:
                t_start = int(tier_match.group(1))
                t_end = int(tier_match.group(2))
                for s in range(t_start, t_end + 1):
                    size_price_map[s] = tier['price']
            else:
                # 单一尺码段，所有尺码同价
                for s in all_sizes:
                    size_price_map[s] = tier['price']

        # 输出格式化的尺码范围（不带#）
        display_range = f"{range_start}-{range_end}"

        # 每个颜色 × 每个尺码 = 一行，每个颜色的第一行插入图片
        for color in colors:
            color_start_row = row
            for size in all_sizes:
                price = size_price_map.get(size, price_tiers[0]['price'] if price_tiers else 0)

                # Col A: PHOTO (border only, image added below)
                ws.cell(row, 1).border = border
                # Col B: STYLE NAME
                ws.cell(row, 2, full_name).font = normal_font
                ws.cell(row, 2).border = border
                ws.cell(row, 2).alignment = left_align
                # Col C: STYLE NO.
                ws.cell(row, 3, style_no).font = normal_font
                ws.cell(row, 3).border = border
                ws.cell(row, 3).alignment = center
                # Col D: COLOR CODE
                ws.cell(row, 4, color).font = normal_font
                ws.cell(row, 4).border = border
                ws.cell(row, 4).alignment = left_align
                # Col E: SIZE RANGE
                ws.cell(row, 5, display_range).font = normal_font
                ws.cell(row, 5).border = border
                ws.cell(row, 5).alignment = center
                # Col F: SIZE
                ws.cell(row, 6, size).font = normal_font
                ws.cell(row, 6).border = border
                ws.cell(row, 6).alignment = center
                # Col G: COG
                ws.cell(row, 7, price).font = normal_font
                ws.cell(row, 7).border = border
                ws.cell(row, 7).alignment = center
                # Col H: ADDITIONAL COMMENT
                if remark:
                    ws.cell(row, 8, remark).font = normal_font
                ws.cell(row, 8).border = border
                ws.cell(row, 8).alignment = left_align

                row += 1

            # 在每个颜色块的第一行插入对应颜色的图片
            color_img = color_images.get(color, image_path)
            if color_img and os.path.exists(color_img):
                try:
                    thumb = _make_thumbnail(color_img, THUMB_W, THUMB_H)
                    if thumb and os.path.exists(thumb):
                        img = XlImage(thumb)
                        ws.add_image(img, f'A{color_start_row}')
                        ws.row_dimensions[color_start_row].height = IMG_ROW_HEIGHT
                except Exception:
                    pass

    _apply_notes_to_workbook(wb, output_path, notes, 'COG Overview')
    return row - 7  # 返回数据行数


# ============================================================
# 路由
# ============================================================
@app.route('/')
def index():
    cleanup_uploads()
    output_files = []
    output_dir = app.config['OUTPUT_FOLDER']
    if os.path.exists(output_dir):
        for f in sorted(os.listdir(output_dir), reverse=True):
            if not f.startswith('.'):
                fpath = os.path.join(output_dir, f)
                size_kb = os.path.getsize(fpath) / 1024
                mtime = datetime.fromtimestamp(os.path.getmtime(fpath)).strftime('%Y-%m-%d %H:%M')
                output_files.append({'name': f, 'size': f'{size_kb:.1f} KB', 'time': mtime})
    return render_template('index.html', output_files=output_files)


@app.route('/pi2po', methods=['POST'])
def pi_to_po():
    if 'pi_file' not in request.files:
        flash('请上传PO文件', 'error')
        return redirect(url_for('index'))

    file = request.files['pi_file']
    if not file.filename:
        flash('请选择文件', 'error')
        return redirect(url_for('index'))

    pi_format = request.form.get('pi_format', 'PR').strip().upper()

    filename = safe_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)

    try:
        if filename.lower().endswith(('.xlsx', '.xls')):
            pi_data = parse_pi_excel(filepath)
        elif filename.lower().endswith('.pdf'):
            pi_data = parse_pi_pdf(filepath)
        elif filename.lower().endswith(('.jpg', '.jpeg', '.png', '.webp', '.heic')):
            pi_data = parse_pi_image([filepath])
        else:
            flash('不支持的文件格式，请上传 Excel (.xlsx/.xls)、PDF 或图片', 'error')
            return redirect(url_for('index'))

        if not pi_data or not pi_data.get('items'):
            flash('无法解析PO文件内容，请检查文件格式是否正确', 'error')
            return redirect(url_for('index'))

        order_id = pi_data.get('invoice_no') or pi_data.get('order_no') or 'PO'
        order_id = re.sub(r'[/\\:*?"<>|]', '-', order_id)  # 清理文件名非法字符
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_name = f"PO_{pi_format}_{order_id}_{timestamp}.xlsx"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_name)

        notes = request.form.get('notes', '').strip()
        generate_po_excel(pi_data, output_path, pi_format=pi_format, notes=notes)

        currency = pi_data.get('currency') or 'USD'
        total_pcs = pi_data.get('total_pieces') or 0
        total_amt = pi_data.get('total_amount') or 0
        flash(f'PO 已生成：{output_name}（共 {int(total_pcs):,} 双，{currency} {float(total_amt):,.2f}）', 'success')

    except Exception as e:
        flash(f'处理出错：{str(e)}', 'error')
        traceback.print_exc()

    return redirect(url_for('index'))


@app.route('/po2packing', methods=['POST'])
def po_to_packing():
    if 'po_file' not in request.files:
        flash('请上传PO文件', 'error')
        return redirect(url_for('index'))

    file = request.files['po_file']
    if not file.filename:
        flash('请选择文件', 'error')
        return redirect(url_for('index'))

    fty_order = request.form.get('fty_order', '').strip()

    filename = safe_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)

    try:
        if filename.lower().endswith(('.xlsx', '.xls')):
            pi_data = parse_pi_excel(filepath)
        elif filename.lower().endswith('.pdf'):
            pi_data = parse_pi_pdf(filepath)
        elif filename.lower().endswith(('.jpg', '.jpeg', '.png', '.webp', '.heic')):
            pi_data = parse_pi_image([filepath])
        else:
            flash('不支持的文件格式，请上传 Excel (.xlsx/.xls)、PDF 或图片', 'error')
            return redirect(url_for('index'))

        if not pi_data or not pi_data.get('items'):
            flash('无法解析PO文件内容，请检查文件格式是否正确', 'error')
            return redirect(url_for('index'))

        order_id = pi_data.get('invoice_no') or pi_data.get('order_no') or 'PO'
        order_id = re.sub(r'[/\\:*?"<>|]', '-', str(order_id))
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_name = f"PackingList_{order_id}_{timestamp}.xlsx"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_name)

        notes = request.form.get('notes', '').strip()
        generate_packing_list_excel(pi_data, output_path, fty_order=fty_order, notes=notes)

        total_pcs = pi_data.get('total_pieces') or 0
        item_count = len(pi_data.get('items', []))
        flash(f'Packing List 已生成：{output_name}（{item_count} 个款式/颜色，共 {int(total_pcs):,} 双）', 'success')

    except Exception as e:
        flash(f'处理出错：{str(e)}', 'error')
        traceback.print_exc()

    return redirect(url_for('index'))


@app.route('/pi2ci', methods=['POST'])
def pi_to_ci():
    if 'ci_file' not in request.files:
        flash('请上传PI文件', 'error')
        return redirect(url_for('index'))

    file = request.files['ci_file']
    if not file.filename:
        flash('请选择文件', 'error')
        return redirect(url_for('index'))

    filename = safe_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)

    try:
        if not filename.lower().endswith(('.xlsx', '.xls')):
            flash('CI转换仅支持 Excel (.xlsx/.xls) 格式的PI文件', 'error')
            return redirect(url_for('index'))

        # Extract order number from filename for output naming
        order_id = re.sub(r'\.(xlsx|xls)$', '', filename, flags=re.IGNORECASE)
        order_id = re.sub(r'^(PI|pi)', '', order_id).strip('_- ')
        if not order_id:
            order_id = 'ORDER'
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_name = f"CI_{order_id}_{timestamp}.xlsx"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_name)

        notes = request.form.get('notes', '').strip()
        generate_ci_from_pi(filepath, output_path, notes=notes)

        flash(f'CI 已生成：{output_name}', 'success')

    except Exception as e:
        flash(f'处理出错：{str(e)}', 'error')
        traceback.print_exc()

    return redirect(url_for('index'))


@app.route('/materials2production', methods=['POST'])
def materials_to_production():
    files = request.files.getlist('material_files')
    if not files or not files[0].filename:
        flash('请上传原材料照片', 'error')
        return redirect(url_for('index'))

    image_paths = []
    for f in files:
        if f.filename:
            safe_name = f"{datetime.now().strftime('%H%M%S')}_{safe_filename(f.filename)}"
            fpath = os.path.join(app.config['UPLOAD_FOLDER'], safe_name)
            f.save(fpath)
            image_paths.append(fpath)

    try:
        material_data = parse_handwritten_materials(image_paths)
        if not material_data:
            flash('无法识别手写内容，请确保照片清晰', 'error')
            return redirect(url_for('index'))

        order_id = material_data.get('order_no', 'ORDER')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        notes = request.form.get('notes', '').strip()
        html_content = generate_production_sheet_html(material_data, notes=notes)

        html_name = f"Production_{order_id}_{timestamp}.html"
        html_path = os.path.join(app.config['OUTPUT_FOLDER'], html_name)
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

        mat_count = len(material_data.get('materials', []))
        hw_count = len(material_data.get('hardware', []))
        flash(f'生产指令单已生成：{html_name}（{mat_count} 种原材料，{hw_count} 种五金）— 点击预览后可打印为PDF', 'success')

    except Exception as e:
        flash(f'处理出错：{str(e)}', 'error')
        traceback.print_exc()

    return redirect(url_for('index'))


@app.route('/quotation2cog', methods=['POST'])
def quotation_to_cog():
    if 'quotation_file' not in request.files:
        flash('请上传报价单文件', 'error')
        return redirect(url_for('index'))

    file = request.files['quotation_file']
    if not file.filename:
        flash('请选择文件', 'error')
        return redirect(url_for('index'))

    filename = safe_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)

    try:
        quotation_data = parse_quotation(filepath)
        if not quotation_data or not quotation_data.get('entries'):
            flash('无法解析报价单，请检查文件格式', 'error')
            return redirect(url_for('index'))

        season = quotation_data.get('season', 'SS')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_name = f"{season}_SMS_COG_OVERVIEW--EMILY_{timestamp}.xlsx"
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_name)

        brand_prefix = request.form.get('brand_prefix', 'bisgaard').strip()
        notes = request.form.get('notes', '').strip()
        total_rows = generate_cog_excel(quotation_data, output_path, brand_prefix=brand_prefix, notes=notes)

        entry_count = len(quotation_data['entries'])
        flash(f'COG Overview 已生成：{output_name}（{entry_count} 个款式，{total_rows} 行数据）', 'success')

    except Exception as e:
        flash(f'处理出错：{str(e)}', 'error')
        traceback.print_exc()

    return redirect(url_for('index'))


@app.route('/download/<filename>')
def download(filename):
    filename = safe_filename(filename)
    filepath = os.path.join(app.config['OUTPUT_FOLDER'], filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    flash('文件不存在', 'error')
    return redirect(url_for('index'))


@app.route('/preview/<filename>')
def preview(filename):
    filename = safe_filename(filename)
    filepath = os.path.join(app.config['OUTPUT_FOLDER'], filename)
    if os.path.exists(filepath) and filepath.endswith('.html'):
        with open(filepath, 'r', encoding='utf-8') as f:
            return f.read()
    flash('文件不存在', 'error')
    return redirect(url_for('index'))


@app.route('/delete/<filename>', methods=['POST'])
def delete_file(filename):
    filename = safe_filename(filename)
    filepath = os.path.join(app.config['OUTPUT_FOLDER'], filename)
    if os.path.exists(filepath):
        os.remove(filepath)
        flash(f'已删除：{filename}', 'success')
    return redirect(url_for('index'))


@app.route('/clear-all', methods=['POST'])
def clear_all():
    """Clear all output files"""
    output_dir = app.config['OUTPUT_FOLDER']
    count = 0
    for f in os.listdir(output_dir):
        fpath = os.path.join(output_dir, f)
        if os.path.isfile(fpath) and not f.startswith('.'):
            os.remove(fpath)
            count += 1
    flash(f'已清空 {count} 个文件', 'success')
    return redirect(url_for('index'))


FEEDBACK_IMG_DIR = os.path.join(BASE_DIR, 'feedback_images')
FEEDBACK_FILE_DIR = os.path.join(BASE_DIR, 'feedback_files')
os.makedirs(FEEDBACK_IMG_DIR, exist_ok=True)
os.makedirs(FEEDBACK_FILE_DIR, exist_ok=True)


@app.route('/feedback', methods=['POST'])
def submit_feedback():
    """Save user feedback with optional images"""
    data = request.get_json(silent=True) or {}
    msg = (data.get('message') or '').strip()
    category = data.get('category', 'general').strip()
    images = data.get('images') or []
    if not msg and not images:
        return jsonify({'ok': False, 'error': '反馈内容为空'}), 400
    if len(msg) > 2000:
        return jsonify({'ok': False, 'error': '反馈内容过长'}), 400
    if len(images) > 5:
        images = images[:5]
    # Load existing
    feedbacks = []
    if os.path.exists(FEEDBACK_FILE):
        try:
            with open(FEEDBACK_FILE, 'r', encoding='utf-8') as f:
                feedbacks = json.load(f)
        except Exception:
            feedbacks = []
    fb_id = len(feedbacks) + 1
    # Save images
    saved_images = []
    for i, img in enumerate(images):
        img_data = img.get('data', '')
        if not img_data or ',' not in img_data:
            continue
        header, b64 = img_data.split(',', 1)
        ext = 'jpg'
        if 'png' in header:
            ext = 'png'
        fname = f'fb_{fb_id}_{i+1}.{ext}'
        fpath = os.path.join(FEEDBACK_IMG_DIR, fname)
        try:
            with open(fpath, 'wb') as f:
                f.write(base64.b64decode(b64))
            saved_images.append(fname)
        except Exception:
            pass
    # Save attached files
    attached_files = data.get('files') or []
    if len(attached_files) > 3:
        attached_files = attached_files[:3]
    ALLOWED_EXT = {'.xlsx', '.xls', '.pdf', '.csv', '.json', '.txt'}
    saved_files = []
    for i, af in enumerate(attached_files):
        af_data = af.get('data', '')
        af_name = af.get('name', f'file_{i}')
        ext = os.path.splitext(af_name)[1].lower()
        if ext not in ALLOWED_EXT or ',' not in af_data:
            continue
        _, b64 = af_data.split(',', 1)
        safe_name = f'fb_{fb_id}_{i+1}{ext}'
        fpath = os.path.join(FEEDBACK_FILE_DIR, safe_name)
        try:
            with open(fpath, 'wb') as f:
                f.write(base64.b64decode(b64))
            saved_files.append({'name': af_name, 'path': safe_name})
        except Exception:
            pass
    feedbacks.append({
        'id': fb_id,
        'message': msg,
        'category': category,
        'images': saved_images,
        'files': saved_files,
        'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'status': 'pending',
    })
    with open(FEEDBACK_FILE, 'w', encoding='utf-8') as f:
        json.dump(feedbacks, f, ensure_ascii=False, indent=2)
    return jsonify({'ok': True})


@app.route('/feedback', methods=['GET'])
def get_feedback():
    """Read all feedback (for Claude Code monitoring)"""
    if os.path.exists(FEEDBACK_FILE):
        with open(FEEDBACK_FILE, 'r', encoding='utf-8') as f:
            return jsonify(json.load(f))
    return jsonify([])


@app.route('/feedback/image/<filename>')
def feedback_image(filename):
    """Serve feedback images"""
    safe = os.path.basename(filename)
    fpath = os.path.join(FEEDBACK_IMG_DIR, safe)
    if os.path.isfile(fpath):
        return send_file(fpath)
    return 'Not found', 404


@app.route('/feedback/file/<filename>')
def feedback_file(filename):
    """Serve feedback attached files"""
    safe = os.path.basename(filename)
    fpath = os.path.join(FEEDBACK_FILE_DIR, safe)
    if os.path.isfile(fpath):
        return send_file(fpath, as_attachment=True)
    return 'Not found', 404


@app.route('/api/health')
def health():
    return jsonify({
        'status': 'ok',
        'version': '2.2',
        'ai_mode': 'api' if ANTHROPIC_API_KEY else 'cli',
        'timestamp': datetime.now().isoformat(),
    })


if __name__ == '__main__':
    port = int(sys.argv[1]) if len(sys.argv) > 1 else int(os.environ.get('PORT', 5566))
    debug = os.environ.get('FLASK_ENV') != 'production'
    print(f"\n{'='*50}")
    print(f"  EMILY 订单管理系统 v2.1")
    print(f"  地址: http://0.0.0.0:{port}")
    print(f"  AI模式: {'API (Anthropic)' if ANTHROPIC_API_KEY else 'CLI (Claude Code)'}")
    print(f"  环境: {'开发' if debug else '生产'}")
    print(f"{'='*50}\n")
    app.run(host='0.0.0.0', port=port, debug=debug)
